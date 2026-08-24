"""Export completo de ficha de contacto + históricos (Excel multihoja y PDF)."""
from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from typing import Any

from services.history_service import HISTORY_SPECS, HistoryKind

HEADER_FILL = "2D6A4F"
HEADER_FONT = "FFFFFF"
MAX_COL_WIDTH = 50
PDF_CELL_MAX_CHARS = 500

# Orden fijo de históricos en el export (LLM-friendly).
EXPORT_HISTORY_KINDS: tuple[HistoryKind, ...] = (
    "seguimiento_comercial",
    "tareas",
    "notas",
    "sensores",
    "campanas",
    "suscripciones",
    "incidencias",
)

SHEET_TITLES: dict[str, str] = {
    "seguimiento_comercial": "Seguimiento",
    "tareas": "Tareas",
    "notas": "Notas",
    "sensores": "Sensores",
    "campanas": "Campanas",
    "suscripciones": "Suscripciones",
    "incidencias": "Incidencias",
}


def _slug_nombre(nombre: str) -> str:
    raw = (nombre or "contacto").strip().lower()
    raw = re.sub(r"[^a-z0-9áéíóúñü]+", "_", raw, flags=re.IGNORECASE)
    raw = re.sub(r"_+", "_", raw).strip("_")
    return (raw or "contacto")[:40]


def _export_filename(contact: dict[str, str], ext: str, exported_at: datetime) -> str:
    nombre = str(contact.get("nombre", "") or "").strip() or "contacto"
    cid = str(contact.get("contact_id", "") or "").strip()[:8]
    slug = _slug_nombre(nombre)
    suffix = f"_{cid}" if cid else ""
    return f"contacto_{slug}{suffix}_{exported_at.strftime('%Y%m%d')}.{ext}"


def _nonempty_ficha_fields(contact: dict[str, str]) -> list[tuple[str, str]]:
    """Campo/valor de la ficha; siempre incluye contact_id si existe."""
    rows: list[tuple[str, str]] = []
    cid = str(contact.get("contact_id", "") or "").strip()
    if cid:
        rows.append(("contact_id", cid))
    for key in sorted(contact.keys()):
        if key == "contact_id":
            continue
        value = str(contact.get(key, "") or "").strip()
        if value:
            rows.append((key, value))
    return rows


def _history_headers(kind: HistoryKind) -> list[str]:
    return list(HISTORY_SPECS[kind].headers)


def _history_rows(kind: HistoryKind, rows: list[dict[str, str]]) -> list[list[str]]:
    headers = _history_headers(kind)
    out: list[list[str]] = []
    for row in rows:
        out.append([str(row.get(h, "") or "") for h in headers])
    return out


def collect_histories_for_contact(history_service: Any, contact_id: str) -> dict[str, list[dict[str, str]]]:
    """Carga los históricos exportables de un contacto."""
    cid = str(contact_id or "").strip()
    result: dict[str, list[dict[str, str]]] = {}
    for kind in EXPORT_HISTORY_KINDS:
        result[kind] = list(history_service.rows_for_contact(kind, cid)) if cid else []
    return result


def build_contact_ficha_xlsx_bytes(
    contact: dict[str, str],
    histories_by_kind: dict[str, list[dict[str, str]]],
    exported_at: datetime | None = None,
) -> tuple[bytes, str]:
    """Return (xlsx_bytes, filename)."""
    now = exported_at or datetime.now()
    filename = _export_filename(contact, "xlsx", now)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color=HEADER_FONT)

    # --- Indice ---
    ws_idx = wb.active
    ws_idx.title = "Indice"
    meta = [
        ("Documento", "Exportación CRM Sanzar — ficha completa de contacto"),
        ("contact_id", str(contact.get("contact_id", "") or "")),
        ("nombre", str(contact.get("nombre", "") or "")),
        ("generado_el", now.strftime("%d/%m/%Y %H:%M")),
        (
            "hojas",
            "Indice, Ficha, "
            + ", ".join(SHEET_TITLES[k] for k in EXPORT_HISTORY_KINDS),
        ),
        (
            "instruccion_llm",
            "Este archivo contiene la ficha del cliente y todos sus históricos. "
            "Lee la hoja Ficha y cada hoja de histórico para tener el contexto completo.",
        ),
    ]
    ws_idx["A1"] = "Campo"
    ws_idx["B1"] = "Valor"
    for col in (1, 2):
        cell = ws_idx.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    for i, (k, v) in enumerate(meta, start=2):
        ws_idx.cell(row=i, column=1, value=k).border = border
        ws_idx.cell(row=i, column=2, value=v).border = border
    ws_idx.column_dimensions["A"].width = 22
    ws_idx.column_dimensions["B"].width = 80

    # --- Ficha ---
    ws_ficha = wb.create_sheet("Ficha")
    ws_ficha["A1"] = "Campo"
    ws_ficha["B1"] = "Valor"
    for col in (1, 2):
        cell = ws_ficha.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    ficha_rows = _nonempty_ficha_fields(contact)
    if not ficha_rows:
        ws_ficha.cell(row=2, column=1, value="(sin datos)").border = border
        ws_ficha.cell(row=2, column=2, value="").border = border
    else:
        for i, (k, v) in enumerate(ficha_rows, start=2):
            ws_ficha.cell(row=i, column=1, value=k).border = border
            cell_v = ws_ficha.cell(row=i, column=2, value=v)
            cell_v.border = border
            cell_v.alignment = Alignment(wrap_text=True, vertical="top")
    ws_ficha.column_dimensions["A"].width = 28
    ws_ficha.column_dimensions["B"].width = 60

    # --- Históricos ---
    for kind in EXPORT_HISTORY_KINDS:
        title = SHEET_TITLES[kind]
        headers = _history_headers(kind)
        rows = _history_rows(kind, histories_by_kind.get(kind, []))
        ws = wb.create_sheet(title)
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, horizontal="center")
        if not rows:
            note = ws.cell(row=2, column=1, value="Sin registros")
            note.border = border
        else:
            for r_idx, row_vals in enumerate(rows, start=2):
                for c_idx, val in enumerate(row_vals, start=1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=val)
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        for col_idx, header in enumerate(headers, start=1):
            max_len = len(header)
            for r_idx in range(2, 2 + max(len(rows), 1)):
                val = ws.cell(row=r_idx, column=col_idx).value
                max_len = max(max_len, min(len(str(val or "")), MAX_COL_WIDTH))
            letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[letter].width = min(max_len + 2, MAX_COL_WIDTH)
        ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), filename


def _escape_xml(s: str) -> str:
    return (
        (s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _pdf_cell_text(value: str) -> str:
    text = str(value or "").strip()
    if len(text) > PDF_CELL_MAX_CHARS:
        return text[: PDF_CELL_MAX_CHARS - 1] + "…"
    return text


def build_contact_ficha_pdf_bytes(
    contact: dict[str, str],
    histories_by_kind: dict[str, list[dict[str, str]]],
    exported_at: datetime | None = None,
) -> tuple[bytes, str]:
    """Return (pdf_bytes, filename). Portrait A4, LongTable for histories."""
    now = exported_at or datetime.now()
    filename = _export_filename(contact, "pdf", now)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            LongTable,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            TableStyle,
        )
    except ModuleNotFoundError:
        minimal = (
            "%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
            "2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
            "3 0 obj<</Type/Page/MediaBox[0 0 612 792]>>endobj\n"
            "trailer<</Root 1 0 R>>\n%%EOF\n"
        )
        return minimal.encode("latin-1"), filename

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="FichaExportTitle",
        parent=styles["Heading1"],
        fontSize=14,
        spaceAfter=6,
    )
    h2_style = ParagraphStyle(
        name="FichaExportH2",
        parent=styles["Heading2"],
        fontSize=11,
        spaceBefore=10,
        spaceAfter=4,
    )
    body_style = ParagraphStyle(
        name="FichaExportBody",
        parent=styles["Normal"],
        fontSize=9,
        leading=11,
    )
    cell_style = ParagraphStyle(
        name="FichaExportCell",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
    )
    header_cell_style = ParagraphStyle(
        name="FichaExportHeaderCell",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
        textColor=colors.white,
    )

    nombre = str(contact.get("nombre", "") or "Sin nombre").strip()
    cid = str(contact.get("contact_id", "") or "").strip()

    story: list = []
    story.append(Paragraph(_escape_xml("Exportación CRM — ficha de contacto"), title_style))
    story.append(
        Paragraph(
            _escape_xml(
                f"Contacto: {nombre} | contact_id: {cid or 'n/d'} | "
                f"Generado: {now.strftime('%d/%m/%Y %H:%M')}"
            ),
            body_style,
        )
    )
    story.append(
        Paragraph(
            _escape_xml(
                "Documento de contexto completo (ficha + históricos) para lectura humana o LLM."
            ),
            body_style,
        )
    )
    story.append(Spacer(1, 0.3 * cm))

    story.append(Paragraph(_escape_xml("1. Ficha del cliente"), h2_style))
    ficha_rows = _nonempty_ficha_fields(contact)
    if not ficha_rows:
        story.append(Paragraph(_escape_xml("Sin datos en la ficha."), body_style))
    else:
        for key, value in ficha_rows:
            story.append(
                Paragraph(
                    f"<b>{_escape_xml(key)}:</b> {_escape_xml(_pdf_cell_text(value))}",
                    body_style,
                )
            )

    section_num = 2
    for kind in EXPORT_HISTORY_KINDS:
        title = SHEET_TITLES[kind]
        spec_title = HISTORY_SPECS[kind].title
        story.append(
            Paragraph(
                _escape_xml(f"{section_num}. {spec_title} ({title})"),
                h2_style,
            )
        )
        section_num += 1
        headers = _history_headers(kind)
        rows = histories_by_kind.get(kind, [])
        if not rows:
            story.append(Paragraph(_escape_xml("Sin registros."), body_style))
            continue

        # Limitar columnas en PDF a summary + id + fechas clave para legibilidad;
        # Excel lleva todas. Usamos headers completos pero truncados en celda.
        # Si hay muchas columnas, usar summary_columns + id_column.
        summary = list(HISTORY_SPECS[kind].summary_columns)
        id_col = HISTORY_SPECS[kind].id_column
        pdf_cols = [id_col] + [c for c in summary if c != id_col]
        # Añadir un par de campos útiles si caben
        for extra in headers:
            if extra not in pdf_cols and len(pdf_cols) < 6:
                if extra in {"contact_id", "nombre_cliente", "created_at", "updated_at"}:
                    continue
                pdf_cols.append(extra)

        n_cols = len(pdf_cols)
        col_w = doc.width / max(n_cols, 1)
        tbl_data: list[list[object]] = [
            [Paragraph(_escape_xml(h), header_cell_style) for h in pdf_cols]
        ]
        for row in rows:
            tbl_data.append(
                [
                    Paragraph(
                        _escape_xml(_pdf_cell_text(str(row.get(h, "") or ""))),
                        cell_style,
                    )
                    for h in pdf_cols
                ]
            )

        tbl = LongTable(tbl_data, colWidths=[col_w] * n_cols, repeatRows=1)
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FILL}")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                    ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 3),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 3),
                    ("TOPPADDING", (0, 0), (-1, -1), 3),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
                ]
            )
        )
        story.append(tbl)

    doc.build(story)
    return buf.getvalue(), filename
