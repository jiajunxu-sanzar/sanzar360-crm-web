"""Motor de umbrales de riego: sensor (CC/PMP), Open-Meteo, Kc FAO-56, informe y Excel."""

from __future__ import annotations

import contextlib
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Optional

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import requests
from matplotlib.figure import Figure



# ===== SENSOR / CC / PMP VWC =====

TABLA_TEXTURAS = {
    "arena":                   {"cc_teorica": 10, "pmp_teorica": 5},
    "arenoso-franco":          {"cc_teorica": 12, "pmp_teorica": 5},
    "franco-arenoso":          {"cc_teorica": 18, "pmp_teorica": 8},
    "franco-arcillo-arenoso":  {"cc_teorica": 27, "pmp_teorica": 17},
    "franco":                  {"cc_teorica": 28, "pmp_teorica": 14},
    "arcillo-arenoso":         {"cc_teorica": 36, "pmp_teorica": 25},
    "franco-limoso":           {"cc_teorica": 31, "pmp_teorica": 11},
    "limo":                    {"cc_teorica": 30, "pmp_teorica": 6},
    "franco-arcilloso":        {"cc_teorica": 36, "pmp_teorica": 22},
    "franco-arcillo-limoso":   {"cc_teorica": 38, "pmp_teorica": 22},
    "arcillo-limoso":          {"cc_teorica": 41, "pmp_teorica": 27},
    "arcilla":                 {"cc_teorica": 42, "pmp_teorica": 30},
}


@dataclass
class ParametrosDeteccion:
    """Umbrales de deteccion, configurables por sensor/cultivo."""

    # Deteccion automatica de riegos (si no se aporta un log de riegos)
    umbral_subida_vwc_h: float = 0.3      # %VWC/h -> por encima, se considera "subida por riego/lluvia"
    horas_lookahead_pico: float = 6.0     # horas tras el inicio de subida en las que se busca el pico

    # Deteccion de estabilizacion (drenaje completado)
    ventana_suavizado: int = 3            # nº de lecturas para la media movil de suavizado
    umbral_estable_vwc_h: float = 0.05    # %VWC/h -> por debajo, se considera "estable"
    horas_min_estable: float = 4.0        # duracion minima sostenida por debajo del umbral

    # Validez del evento
    horas_min_ventana_drenaje: float = 24.0  # ventana minima disponible tras el pico para aceptar el evento
    horas_max_busqueda_estabilizacion: float = 96.0  # tope de busqueda de estabilizacion

    # Robustez del agregado final
    n_min_eventos: int = 5                # nº minimo de eventos validos para considerar el resultado robusto


@dataclass
class ResultadoEvento:
    inicio_riego: pd.Timestamp
    pico_tiempo: pd.Timestamp
    pico_valor: float
    estabilizacion_tiempo: Optional[pd.Timestamp]
    cc_evento: Optional[float]
    valido: bool
    motivo_descarte: Optional[str] = None


@dataclass
class ResultadoUmbrales:
    umbral_superior_cc_optima: Optional[float]
    umbral_inferior_raw: Optional[float]
    ad: Optional[float]
    pmp_teorico: float
    p: float
    n_eventos_detectados: int
    n_eventos_validos: int
    robusto: bool
    eventos: list = field(default_factory=list)


# ----------------------------------------------------------------------
# Carga de datos
# ----------------------------------------------------------------------
def _ensure_datetime64_ns(values) -> pd.Series:
    """Normaliza timestamps a datetime64[ns].

    En Python 3.12+ / pandas reciente, ``pd.to_datetime`` puede devolver
    ``datetime64[us]``. ``merge_asof`` exige la misma unidad en ambas
    claves (ns vs us → MergeError). Unificamos a ns.
    """
    series = pd.to_datetime(values)
    if isinstance(series, pd.Series):
        try:
            return series.astype("datetime64[ns]")
        except (TypeError, ValueError):
            if hasattr(series.dt, "as_unit"):
                return series.dt.as_unit("ns")
            return series
    # Index u otro array-like
    series = pd.Series(series)
    try:
        return series.astype("datetime64[ns]")
    except (TypeError, ValueError):
        if hasattr(series.dt, "as_unit"):
            return series.dt.as_unit("ns")
        return series


def cargar_serie(csv_path: str, col_timestamp: int = 5, col_valor: int = 6,
                  con_cabecera: bool = False) -> pd.DataFrame:
    """
    Carga el CSV del sensor y devuelve un DataFrame con columnas
    ['timestamp', 'valor'] ordenado cronologicamente. Pensado para el
    formato sin cabecera usado por los sensores TEROS10 (columnas por
    posicion), pero admite CSVs con cabecera si con_cabecera=True y las
    columnas ya se llaman 'timestamp' y 'valor'.
    """
    if con_cabecera:
        df = pd.read_csv(csv_path)
        df = df.rename(columns={df.columns[col_timestamp]: "timestamp",
                                 df.columns[col_valor]: "valor"})
    else:
        raw = pd.read_csv(csv_path, header=None)
        df = pd.DataFrame({
            "timestamp": raw[col_timestamp],
            "valor": raw[col_valor],
        })

    df["timestamp"] = _ensure_datetime64_ns(df["timestamp"])
    df = df[["timestamp", "valor"]].dropna().sort_values("timestamp").reset_index(drop=True)
    return df


# ----------------------------------------------------------------------
# Deteccion de eventos de riego
# ----------------------------------------------------------------------
def derivada_horaria(df: pd.DataFrame, ventana_suavizado: int) -> pd.Series:
    """Derivada suavizada de 'valor' respecto al tiempo, en %VWC/h."""
    valor_suave = df["valor"].rolling(ventana_suavizado, min_periods=1, center=True).mean()
    horas = df["timestamp"].diff().dt.total_seconds() / 3600.0
    deriv = valor_suave.diff() / horas
    return deriv


def detectar_inicios_riego_automatico(df: pd.DataFrame, params: ParametrosDeteccion) -> list[pd.Timestamp]:
    """
    Detecta automaticamente los inicios de subida (riego o lluvia) cuando
    no se dispone de un registro de riegos. Agrupa tramos consecutivos de
    subida rapida en un unico evento (se queda con el primer instante de
    cada tramo).

    Limitacion conocida: no distingue riego de lluvia. Si se dispone de
    datos de pluviometria o de un registro de riegos, se recomienda
    pasarlos via `inicios_riego_manual` en `calcular_umbrales` en vez de
    usar la deteccion automatica.
    """
    deriv = derivada_horaria(df, params.ventana_suavizado)
    subiendo = deriv > params.umbral_subida_vwc_h

    inicios = []
    en_evento = False
    for i, es_subida in enumerate(subiendo):
        if es_subida and not en_evento:
            inicios.append(df["timestamp"].iloc[i])
            en_evento = True
        elif not es_subida:
            en_evento = False
    return inicios


def _localizar_pico(df: pd.DataFrame, inicio: pd.Timestamp, params: ParametrosDeteccion):
    ventana = df[(df["timestamp"] >= inicio) &
                 (df["timestamp"] <= inicio + pd.Timedelta(hours=params.horas_lookahead_pico))]
    if ventana.empty:
        return None, None
    idx_max = ventana["valor"].idxmax()
    return df["timestamp"].loc[idx_max], df["valor"].loc[idx_max]


def _localizar_estabilizacion(df: pd.DataFrame, pico_tiempo: pd.Timestamp,
                               fin_ventana: pd.Timestamp, params: ParametrosDeteccion):
    """
    Busca, a partir del pico, el primer instante en que la derivada se
    mantiene por debajo de `umbral_estable_vwc_h` durante al menos
    `horas_min_estable` seguidas. Devuelve el timestamp de estabilizacion
    o None si no se alcanza dentro de la ventana disponible.
    """
    tope = min(fin_ventana, pico_tiempo + pd.Timedelta(hours=params.horas_max_busqueda_estabilizacion))
    tramo = df[(df["timestamp"] >= pico_tiempo) & (df["timestamp"] <= tope)].reset_index(drop=True)
    if len(tramo) < 3:
        return None

    deriv = derivada_horaria(tramo, params.ventana_suavizado).abs()

    for i in range(len(tramo)):
        t0 = tramo["timestamp"].iloc[i]
        t_fin_sostenido = t0 + pd.Timedelta(hours=params.horas_min_estable)
        sub = tramo[(tramo["timestamp"] >= t0) & (tramo["timestamp"] <= t_fin_sostenido)]
        sub_deriv = deriv.loc[sub.index]
        if len(sub) >= 2 and sub_deriv.max() <= params.umbral_estable_vwc_h:
            return t0
    return None


def analizar_eventos(df: pd.DataFrame, inicios_riego: list[pd.Timestamp],
                      params: ParametrosDeteccion) -> list[ResultadoEvento]:
    resultados = []
    inicios_ordenados = sorted(inicios_riego)

    for i, inicio in enumerate(inicios_ordenados):
        siguiente_inicio = inicios_ordenados[i + 1] if i + 1 < len(inicios_ordenados) else df["timestamp"].max()

        pico_tiempo, pico_valor = _localizar_pico(df, inicio, params)
        if pico_tiempo is None:
            resultados.append(ResultadoEvento(inicio, inicio, np.nan, None, None, False, "sin_datos_tras_inicio"))
            continue

        ventana_disponible_h = (siguiente_inicio - pico_tiempo).total_seconds() / 3600.0
        if ventana_disponible_h < params.horas_min_ventana_drenaje:
            resultados.append(ResultadoEvento(inicio, pico_tiempo, pico_valor, None, None, False,
                                               "ventana_drenaje_insuficiente"))
            continue

        t_estable = _localizar_estabilizacion(df, pico_tiempo, siguiente_inicio, params)
        if t_estable is None:
            resultados.append(ResultadoEvento(inicio, pico_tiempo, pico_valor, None, None, False,
                                               "no_estabiliza_en_ventana"))
            continue

        meseta = df[(df["timestamp"] >= t_estable) & (df["timestamp"] <= siguiente_inicio)]
        cc_evento = meseta["valor"].median()

        resultados.append(ResultadoEvento(inicio, pico_tiempo, pico_valor, t_estable, cc_evento, True))

    return resultados


# ----------------------------------------------------------------------
# Calculo final de umbrales
# ----------------------------------------------------------------------
def calcular_umbrales(csv_path: str, p: float, textura: Optional[str] = None,
                       pmp_teorico: Optional[float] = None,
                       inicios_riego_manual: Optional[list[str]] = None,
                       params: Optional[ParametrosDeteccion] = None,
                       col_timestamp: int = 5, col_valor: int = 6,
                       con_cabecera: bool = False) -> ResultadoUmbrales:
    """
    Calcula el umbral superior (CC optima real) y el umbral inferior (RAW)
    a partir de la serie temporal de un sensor.

    Parametros
    ----------
    csv_path : ruta al CSV del sensor (timestamp, valor).
    p : fraccion de agotamiento permitido para el cultivo (FAO-56, Cuadro 22).
        Parametro de entrada obligatorio, no se calcula.
    textura : nombre de la clase textural (ver TABLA_TEXTURAS) para obtener
        el PMP_teorico de tabla. Ignorado si se pasa `pmp_teorico` directamente.
    pmp_teorico : PMP teorico en %VWC, si se quiere dar el valor directamente
        en vez de por textura.
    inicios_riego_manual : lista opcional de timestamps (str) con el registro
        real de riegos. Si se aporta, se usa en vez de la deteccion automatica.
    params : ParametrosDeteccion (opcional); si no se pasa, se usan los
        valores por defecto.
    """
    if pmp_teorico is None:
        if textura is None:
            raise ValueError("Debes indicar 'textura' o 'pmp_teorico'.")
        clave = textura.strip().lower()
        if clave not in TABLA_TEXTURAS:
            raise ValueError(f"Textura '{textura}' no reconocida. Opciones: {list(TABLA_TEXTURAS)}")
        pmp_teorico = TABLA_TEXTURAS[clave]["pmp_teorica"]

    params = params or ParametrosDeteccion()
    df = cargar_serie(csv_path, col_timestamp=col_timestamp, col_valor=col_valor, con_cabecera=con_cabecera)

    if inicios_riego_manual:
        inicios = [pd.to_datetime(t) for t in inicios_riego_manual]
    else:
        inicios = detectar_inicios_riego_automatico(df, params)

    eventos = analizar_eventos(df, inicios, params)
    eventos_validos = [e for e in eventos if e.valido]

    if eventos_validos:
        cc_optima = float(np.median([e.cc_evento for e in eventos_validos]))
        ad = cc_optima - pmp_teorico
        raw = cc_optima - (p * ad)
    else:
        cc_optima = None
        ad = None
        raw = None

    return ResultadoUmbrales(
        umbral_superior_cc_optima=cc_optima,
        umbral_inferior_raw=raw,
        ad=ad,
        pmp_teorico=pmp_teorico,
        p=p,
        n_eventos_detectados=len(eventos),
        n_eventos_validos=len(eventos_validos),
        robusto=len(eventos_validos) >= params.n_min_eventos,
        eventos=eventos,
    )


def imprimir_informe(resultado: ResultadoUmbrales) -> None:
    print("=" * 60)
    print("INFORME DE CALCULO DE UMBRALES DE RIEGO")
    print("=" * 60)
    print(f"Eventos de riego detectados : {resultado.n_eventos_detectados}")
    print(f"Eventos validos (usados)    : {resultado.n_eventos_validos}")
    print(f"Resultado robusto (n >= min): {'SI' if resultado.robusto else 'NO -> ampliar periodo de datos'}")
    print("-" * 60)
    if resultado.umbral_superior_cc_optima is not None:
        print(f"UMBRAL SUPERIOR (CC optima) : {resultado.umbral_superior_cc_optima:.2f} %VWC")
        print(f"PMP teorico                 : {resultado.pmp_teorico:.2f} %VWC")
        print(f"AD (agua disponible)        : {resultado.ad:.2f} %VWC")
        print(f"p (fraccion agotamiento)    : {resultado.p:.2f}")
        print(f"UMBRAL INFERIOR (RAW)       : {resultado.umbral_inferior_raw:.2f} %VWC")
    else:
        print("No se pudo calcular: no hay eventos validos con suficiente ventana de drenaje.")
    print("-" * 60)
    print("Detalle por evento:")
    for e in resultado.eventos:
        if e.valido:
            print(f"  [OK] riego {e.inicio_riego} | pico {e.pico_valor:.2f} en {e.pico_tiempo} "
                  f"| estable desde {e.estabilizacion_tiempo} | CC_evento={e.cc_evento:.2f}")
        else:
            print(f"  [--] riego {e.inicio_riego} | descartado: {e.motivo_descarte}")


# ----------------------------------------------------------------------
# PMP OPERATIVO (proxy de secado / dry-down)
# ----------------------------------------------------------------------
# IMPORTANTE: esto NO es el PMP de laboratorio (-1500 kPa de potencial
# matricial). Es un proxy operativo: el punto en el que, durante un tramo
# suficientemente largo sin riego, el ritmo de perdida de humedad pasa de
# la fase de extraccion "activa" (planta transpirando con normalidad) a un
# ritmo mucho mas lento y sostenido (senal de que la extraccion se esta
# frenando). Para tratarlo como PMP real hay que confirmarlo con una medida
# directa de estres de la planta (camara de presion, marchitez visible
# sostenida) en el mismo periodo. Mientras tanto, sirve como el mejor dato
# disponible para revisar/ajustar el PMP teorico de tabla.
#
# Reutiliza la deteccion de riego/CC ya implementada: para cada evento con
# CC valido, mira si la ventana hasta el siguiente riego es lo bastante
# larga como para intentar ver, mas alla de la meseta de CC, un quiebre de
# pendiente en la fase de secado posterior.
# ----------------------------------------------------------------------
@dataclass
class ParametrosPMP:
    """Umbrales de deteccion del quiebre de pendiente en la fase de secado."""

    horas_min_ventana_dry_down: float = 96.0     # ventana minima tras la meseta CC para intentar el analisis (mas larga que un ciclo de riego habitual)
    horas_referencia_pendiente: float = 48.0     # tramo tras la meseta CC usado para estimar la pendiente de "extraccion activa"
    fraccion_quiebre_pendiente: float = 0.40     # el quiebre se marca cuando la pendiente cae por debajo de esta fraccion de la pendiente de referencia
    horas_min_quiebre_sostenido: float = 24.0    # duracion minima sostenida por debajo del umbral relativo
    pendiente_minima_valida_vwc_h: float = 0.03  # pendiente de referencia minima para considerar el tramo "activo" (evita falsos positivos en tramos ya casi planos)
    horas_max_busqueda_quiebre: float = 240.0    # tope de busqueda del quiebre tras el tramo de referencia
    n_min_eventos_pmp: int = 2                   # eventos minimos para considerar el resultado robusto (los dry-downs largos son mas raros que los riegos)


@dataclass
class ResultadoEventoPMP:
    evento_cc: ResultadoEvento
    ventana_disponible_h: float
    pendiente_referencia: Optional[float]
    quiebre_tiempo: Optional[pd.Timestamp]
    pmp_evento: Optional[float]
    valido: bool
    motivo_descarte: Optional[str] = None


@dataclass
class ResultadoPMPOperativo:
    pmp_operativo: Optional[float]
    n_ventanas_analizadas: int
    n_eventos_validos: int
    robusto: bool
    eventos: list = field(default_factory=list)


def _pendiente_regresion(df_tramo: pd.DataFrame) -> Optional[float]:
    """Pendiente (valor/h) por regresion lineal simple sobre un tramo de datos."""
    if len(df_tramo) < 3:
        return None
    horas = (df_tramo["timestamp"] - df_tramo["timestamp"].iloc[0]).dt.total_seconds() / 3600.0
    pendiente, _ = np.polyfit(horas, df_tramo["valor"], 1)
    return pendiente


def _buscar_quiebre_pendiente(df: pd.DataFrame, inicio_busqueda: pd.Timestamp, fin_ventana: pd.Timestamp,
                               pendiente_referencia: float, params: ParametrosPMP) -> Optional[pd.Timestamp]:
    """
    Analogo a `_localizar_estabilizacion` pero para el quiebre de pendiente
    en la fase de secado: busca el primer instante en que la pendiente
    local se mantiene, en valor absoluto, por debajo de una fraccion de la
    pendiente de referencia (fase de extraccion activa), durante un tiempo
    sostenido.
    """
    tope = min(fin_ventana, inicio_busqueda + pd.Timedelta(hours=params.horas_max_busqueda_quiebre))
    tramo = df[(df["timestamp"] >= inicio_busqueda) & (df["timestamp"] <= tope)].reset_index(drop=True)
    if len(tramo) < 3:
        return None

    deriv = derivada_horaria(tramo, ventana_suavizado=ParametrosDeteccion.ventana_suavizado).abs()
    umbral = params.fraccion_quiebre_pendiente * abs(pendiente_referencia)

    for i in range(len(tramo)):
        t0 = tramo["timestamp"].iloc[i]
        t_fin_sostenido = t0 + pd.Timedelta(hours=params.horas_min_quiebre_sostenido)
        sub = tramo[(tramo["timestamp"] >= t0) & (tramo["timestamp"] <= t_fin_sostenido)]
        sub_deriv = deriv.loc[sub.index]
        if len(sub) >= 2 and sub_deriv.max() <= umbral:
            return t0
    return None


def analizar_pmp_operativo(df: pd.DataFrame, eventos_cc: list[ResultadoEvento],
                            params: ParametrosPMP) -> ResultadoPMPOperativo:
    """
    Para cada evento de riego con CC ya calculada, comprueba si la ventana
    hasta el siguiente riego es lo bastante larga como para intentar ver un
    quiebre de pendiente en la fase de secado posterior a la meseta de CC.
    """
    inicios_ordenados = sorted(e.inicio_riego for e in eventos_cc)
    resultados = []

    for e in eventos_cc:
        if not e.valido or e.estabilizacion_tiempo is None:
            continue

        posteriores = [t for t in inicios_ordenados if t > e.inicio_riego]
        siguiente_inicio = posteriores[0] if posteriores else df["timestamp"].max()

        ventana_h = (siguiente_inicio - e.estabilizacion_tiempo).total_seconds() / 3600.0
        if ventana_h < params.horas_min_ventana_dry_down:
            resultados.append(ResultadoEventoPMP(e, ventana_h, None, None, None, False,
                                                   "ventana_dry_down_insuficiente"))
            continue

        fin_referencia = e.estabilizacion_tiempo + pd.Timedelta(hours=params.horas_referencia_pendiente)
        tramo_referencia = df[(df["timestamp"] >= e.estabilizacion_tiempo) & (df["timestamp"] <= fin_referencia)]
        pendiente_ref = _pendiente_regresion(tramo_referencia)

        if pendiente_ref is None or abs(pendiente_ref) < params.pendiente_minima_valida_vwc_h:
            resultados.append(ResultadoEventoPMP(e, ventana_h, pendiente_ref, None, None, False,
                                                   "pendiente_referencia_insuficiente"))
            continue

        quiebre_tiempo = _buscar_quiebre_pendiente(df, fin_referencia, siguiente_inicio, pendiente_ref, params)
        if quiebre_tiempo is None:
            resultados.append(ResultadoEventoPMP(e, ventana_h, pendiente_ref, None, None, False,
                                                   "sin_quiebre_detectado"))
            continue

        meseta_baja = df[(df["timestamp"] >= quiebre_tiempo) & (df["timestamp"] <= siguiente_inicio)]
        pmp_evento = meseta_baja["valor"].median()
        resultados.append(ResultadoEventoPMP(e, ventana_h, pendiente_ref, quiebre_tiempo, pmp_evento, True))

    validos = [r for r in resultados if r.valido]
    pmp_operativo = float(np.median([r.pmp_evento for r in validos])) if validos else None

    return ResultadoPMPOperativo(
        pmp_operativo=pmp_operativo,
        n_ventanas_analizadas=len(resultados),
        n_eventos_validos=len(validos),
        robusto=len(validos) >= params.n_min_eventos_pmp,
        eventos=resultados,
    )


def calcular_pmp_operativo(csv_path: str, inicios_riego_manual: Optional[list[str]] = None,
                            params_cc: Optional[ParametrosDeteccion] = None,
                            params_pmp: Optional[ParametrosPMP] = None,
                            col_timestamp: int = 5, col_valor: int = 6,
                            con_cabecera: bool = False) -> ResultadoPMPOperativo:
    """
    Punto de entrada: calcula el PMP operativo (proxy) a partir del CSV del
    sensor, reutilizando la deteccion de eventos/CC ya implementada.
    """
    params_cc = params_cc or ParametrosDeteccion()
    params_pmp = params_pmp or ParametrosPMP()

    df = cargar_serie(csv_path, col_timestamp=col_timestamp, col_valor=col_valor, con_cabecera=con_cabecera)

    if inicios_riego_manual:
        inicios = [pd.to_datetime(t) for t in inicios_riego_manual]
    else:
        inicios = detectar_inicios_riego_automatico(df, params_cc)

    eventos_cc = analizar_eventos(df, inicios, params_cc)
    return analizar_pmp_operativo(df, eventos_cc, params_pmp)


def imprimir_informe_pmp(resultado: ResultadoPMPOperativo) -> None:
    print("=" * 60)
    print("INFORME DE BUSQUEDA DE PMP OPERATIVO (proxy, no de laboratorio)")
    print("=" * 60)
    print(f"Ventanas de dry-down suficientemente largas : {resultado.n_ventanas_analizadas}")
    print(f"Quiebres de pendiente detectados (validos)  : {resultado.n_eventos_validos}")
    print(f"Resultado robusto (n >= min)                 : "
          f"{'SI' if resultado.robusto else 'NO -> se necesitan mas dry-downs'}")
    print("-" * 60)
    if resultado.pmp_operativo is not None:
        print(f"PMP OPERATIVO (proxy)  : {resultado.pmp_operativo:.2f} %VWC")
        print("Recuerda: confirmar con una medida directa de estres de la planta")
        print("antes de tratarlo como PMP real en el calculo de RAW.")
    else:
        print("Sin resultado: ningun dry-down del periodo analizado ha sido lo bastante")
        print("largo o profundo como para mostrar un quiebre de pendiente claro.")
    print("-" * 60)
    print("Detalle por evento:")
    for r in resultado.eventos:
        base = f"  riego {r.evento_cc.inicio_riego} (ventana {r.ventana_disponible_h:.1f}h)"
        if r.valido:
            print(f"{base} [OK] pendiente_ref={r.pendiente_referencia:.3f} %VWC/h | "
                  f"quiebre en {r.quiebre_tiempo} | PMP_evento={r.pmp_evento:.2f}")
        else:
            ref = f"{r.pendiente_referencia:.3f}" if r.pendiente_referencia is not None else "n/d"
            print(f"{base} [--] pendiente_ref={ref} | descartado: {r.motivo_descarte}")


# ----------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------

# ===== OPEN-METEO / PMP ET / EXCEL =====

OPEN_METEO_FORECAST_URL = "https://api.open-meteo.com/v1/forecast"
# Para fechas de mas de ~92 dias atras, usar en su lugar:
OPEN_METEO_ARCHIVE_URL = "https://archive-api.open-meteo.com/v1/archive"


# ----------------------------------------------------------------------
# Descarga de datos meteorologicos
# ----------------------------------------------------------------------
def obtener_et_precipitacion(lat: float, lon: float, past_days: int = 30,
                              forecast_days: int = 1,
                              timezone: str = "Europe/Madrid",
                              usar_archivo_historico: bool = False,
                              fecha_inicio: Optional[str] = None,
                              fecha_fin: Optional[str] = None) -> pd.DataFrame:
    """
    Descarga ET0 (FAO Penman-Monteith) y precipitacion horaria de
    Open-Meteo para una ubicacion.

    Modo por defecto (usar_archivo_historico=False): usa el endpoint de
    forecast con `past_days`, que cubre hasta 92 dias hacia atras desde
    hoy incluyendo los datos mas recientes (mejor opcion para cruzar con
    datos de sensores actuales).

    Modo historico (usar_archivo_historico=True, con fecha_inicio y
    fecha_fin en formato 'YYYY-MM-DD'): usa la Historical Weather API
    (ERA5), pensada para periodos antiguos; tiene unos dias de latencia,
    asi que no cubre los ultimos 4-5 dias.
    """
    if usar_archivo_historico:
        if not (fecha_inicio and fecha_fin):
            raise ValueError("Para el modo historico hay que indicar fecha_inicio y fecha_fin (YYYY-MM-DD).")
        url = OPEN_METEO_ARCHIVE_URL
        params = {
            "latitude": lat,
            "longitude": lon,
            "hourly": "et0_fao_evapotranspiration,precipitation",
            "start_date": fecha_inicio,
            "end_date": fecha_fin,
            "timezone": timezone,
        }
    else:
        url = OPEN_METEO_FORECAST_URL
        params = {
            "latitude": lat,
            "longitude": lon,
            "hourly": "et0_fao_evapotranspiration,precipitation",
            "past_days": past_days,
            "forecast_days": forecast_days,
            "timezone": timezone,
        }

    resp = requests.get(url, params=params, timeout=30)
    resp.raise_for_status()
    datos = resp.json()["hourly"]

    df = pd.DataFrame({
        "timestamp": _ensure_datetime64_ns(datos["time"]),
        "et0": datos["et0_fao_evapotranspiration"],
        "precipitacion": datos["precipitation"],
    })
    return df


def _normalize_meteo_column_name(name: object) -> str:
    raw = str(name or "").strip().lower()
    if raw in {"time", "timestamp", "fecha", "datetime"}:
        return "timestamp"
    if "et0" in raw or "evapotranspiration" in raw:
        return "et0"
    if "precip" in raw or raw in {"lluvia", "rain"}:
        return "precipitacion"
    return ""


def cargar_meteo_open_meteo_csv(path: str) -> pd.DataFrame:
    """Carga un CSV hourly de Open-Meteo (export web o limpio) a timestamp/et0/precipitacion.

    Soporta el export con filas de metadata (latitude, longitude, …) antes de la
    cabecera ``time, et0_fao_evapotranspiration (mm), precipitation (mm)``.
    """
    peek = pd.read_csv(path, header=None, dtype=str, nrows=30)
    header_row: int | None = None
    for i, row in peek.iterrows():
        mapped = [_normalize_meteo_column_name(v) for v in row.tolist()]
        if "timestamp" in mapped and "et0" in mapped and "precipitacion" in mapped:
            header_row = int(i)
            break

    if header_row is None:
        df = pd.read_csv(path)
    else:
        df = pd.read_csv(path, skiprows=header_row)

    rename: dict[str, str] = {}
    for col in df.columns:
        canon = _normalize_meteo_column_name(col)
        if canon and canon not in rename.values():
            rename[col] = canon
    df = df.rename(columns=rename)
    missing = [c for c in ("timestamp", "et0", "precipitacion") if c not in df.columns]
    if missing:
        raise ValueError(
            "El CSV de meteo no tiene las columnas esperadas "
            f"(faltan: {', '.join(missing)}). "
            "Usa el export hourly de Open-Meteo con time, ET0 y precipitation."
        )

    out = pd.DataFrame(
        {
            "timestamp": _ensure_datetime64_ns(df["timestamp"]),
            "et0": pd.to_numeric(df["et0"], errors="coerce"),
            "precipitacion": pd.to_numeric(df["precipitacion"], errors="coerce"),
        }
    )
    out = out.dropna(subset=["timestamp"]).sort_values("timestamp").reset_index(drop=True)
    if out.empty:
        raise ValueError("El CSV de meteo no contiene filas horarias válidas.")
    return out


def aviso_solape_meteo_sensor(
    sensor_inicio: pd.Timestamp,
    sensor_fin: pd.Timestamp,
    df_meteo: pd.DataFrame,
) -> str | None:
    """Aviso si la meteo acaba antes (o empieza después) que el sensor; None si OK."""
    if df_meteo is None or df_meteo.empty:
        return None
    meteo_ini = pd.Timestamp(df_meteo["timestamp"].min())
    meteo_fin = pd.Timestamp(df_meteo["timestamp"].max())
    s_ini = pd.Timestamp(sensor_inicio)
    s_fin = pd.Timestamp(sensor_fin)
    # Más de ~12h de hueco al final o al inicio → avisar.
    if meteo_fin + pd.Timedelta(hours=12) < s_fin:
        return (
            f"Meteo cubre hasta {meteo_fin.strftime('%d/%m/%Y')}; "
            f"el sensor llega al {s_fin.strftime('%d/%m/%Y')}. "
            "El tramo sin meteo se ignora para ET/lluvia; CC/PMP por humedad sí usa todo el sensor."
        )
    if meteo_ini > s_ini + pd.Timedelta(hours=12):
        return (
            f"Meteo empieza el {meteo_ini.strftime('%d/%m/%Y')}; "
            f"el sensor desde {s_ini.strftime('%d/%m/%Y')}. "
            "El tramo inicial sin meteo se ignora para ET/lluvia."
        )
    return None


# ----------------------------------------------------------------------
# 1. Riego vs lluvia
# ----------------------------------------------------------------------
def marcar_posible_lluvia(eventos: list[ResultadoEvento], df_meteo: pd.DataFrame,
                           umbral_mm: float = 1.0, ventana_horas: float = 3.0) -> pd.DataFrame:
    """
    Para cada evento de riego detectado automaticamente, comprueba si hubo
    precipitacion relevante (>= umbral_mm) en las horas alrededor de su
    inicio. Devuelve un DataFrame de diagnostico para poder excluir del
    calculo de CC/PMP los eventos que probablemente sean lluvia, no riego
    programado.
    """
    filas = []
    for e in eventos:
        ventana = df_meteo[(df_meteo["timestamp"] >= e.inicio_riego - pd.Timedelta(hours=ventana_horas)) &
                            (df_meteo["timestamp"] <= e.inicio_riego + pd.Timedelta(hours=ventana_horas))]
        precip_total = ventana["precipitacion"].sum()
        filas.append({
            "inicio_riego": e.inicio_riego,
            "precipitacion_mm_ventana": round(float(precip_total), 2),
            "posible_lluvia": bool(precip_total >= umbral_mm),
        })
    return pd.DataFrame(filas)


# ----------------------------------------------------------------------
# 2. Ajuste dinamico de p (FAO-56, ec. 8-4)
# ----------------------------------------------------------------------
def p_ajustado_fao56(p_tabla: float, etc_diaria_mm: float) -> float:
    """
    p_ajustado = p_tabla + 0.04 * (5 - ETc), acotado entre 0.10 y 0.80.
    ETc en mm/dia (ETc = ET0 * Kc; si no se conoce el Kc del cultivo/fase,
    se puede aproximar con ET0 mientras se consigue el Kc real, dejando
    constancia de que es una aproximacion).
    """
    p_adj = p_tabla + 0.04 * (5 - etc_diaria_mm)
    return float(np.clip(p_adj, 0.10, 0.80))


def tabla_p_ajustado_diario(df_meteo: pd.DataFrame, p_tabla: float, kc: float = 1.0) -> pd.DataFrame:
    """Serie diaria de ETc y p_ajustado para revisar como se habria movido el umbral inferior dia a dia."""
    diario = df_meteo.set_index("timestamp").resample("D")["et0"].sum().rename("et0_diaria_mm").reset_index()
    diario["etc_diaria_mm"] = diario["et0_diaria_mm"] * kc
    diario["p_ajustado"] = diario["etc_diaria_mm"].apply(lambda etc: p_ajustado_fao56(p_tabla, etc))
    return diario


# ----------------------------------------------------------------------
# 3. Quiebre de estres normalizado por ET (PMP operativo, version rigurosa)
# ----------------------------------------------------------------------
def analizar_pmp_operativo_normalizado_et(df: pd.DataFrame, df_meteo: pd.DataFrame,
                                           eventos_cc: list[ResultadoEvento],
                                           kc: float = 1.0,
                                           fraccion_quiebre: float = 0.40,
                                           horas_min_quiebre_sostenido: float = 24.0,
                                           horas_min_ventana_dry_down: float = 96.0,
                                           horas_referencia: float = 12.0) -> list[dict]:
    """
    Variante de `umbrales_riego.analizar_pmp_operativo` que compara el
    ritmo real de extraccion (derivada de VWC) contra la demanda de ET
    (et0 * kc) en vez de mirar solo la pendiente absoluta. El quiebre se
    marca cuando el ratio extraccion/demanda_ET cae de forma sostenida por
    debajo de una fraccion del ratio de referencia (calculado en las
    primeras `horas_referencia` del tramo, cuando se asume extraccion sin
    restriccion).

    Nota tecnica: no hace falta conocer la profundidad efectiva del
    sensor para convertir %VWC/h a mm/h de forma absoluta, porque se
    compara el ratio contra su propia referencia (el mismo factor de
    conversion, constante, se cancela en el cociente).
    """
    inicios_ordenados = sorted(e.inicio_riego for e in eventos_cc)
    resultados = []

    left = df.sort_values("timestamp").copy()
    right = df_meteo.sort_values("timestamp").copy()
    left["timestamp"] = _ensure_datetime64_ns(left["timestamp"])
    right["timestamp"] = _ensure_datetime64_ns(right["timestamp"])
    df_merge = pd.merge_asof(
        left, right,
        on="timestamp", direction="nearest", tolerance=pd.Timedelta("1h"),
    )

    for e in eventos_cc:
        if not e.valido or e.estabilizacion_tiempo is None:
            continue

        posteriores = [t for t in inicios_ordenados if t > e.inicio_riego]
        siguiente_inicio = posteriores[0] if posteriores else df["timestamp"].max()

        ventana_h = (siguiente_inicio - e.estabilizacion_tiempo).total_seconds() / 3600.0
        if ventana_h < horas_min_ventana_dry_down:
            resultados.append({"inicio_riego": e.inicio_riego, "valido": False,
                                "motivo_descarte": "ventana_dry_down_insuficiente"})
            continue

        tramo = df_merge[(df_merge["timestamp"] >= e.estabilizacion_tiempo) &
                          (df_merge["timestamp"] <= siguiente_inicio)].reset_index(drop=True)
        if len(tramo) < 6 or tramo["et0"].isna().all():
            resultados.append({"inicio_riego": e.inicio_riego, "valido": False,
                                "motivo_descarte": "sin_datos_et_suficientes"})
            continue

        extraccion = -derivada_horaria(tramo, ventana_suavizado=ParametrosDeteccion.ventana_suavizado)
        demanda_et = tramo["et0"].fillna(0) * kc
        ratio = extraccion / demanda_et.replace(0, np.nan)

        t_ref_fin = tramo["timestamp"].iloc[0] + pd.Timedelta(hours=horas_referencia)
        mask_ref = tramo["timestamp"] <= t_ref_fin
        ratio_referencia = ratio.loc[mask_ref].dropna().median()
        if pd.isna(ratio_referencia) or ratio_referencia <= 0:
            resultados.append({"inicio_riego": e.inicio_riego, "valido": False,
                                "motivo_descarte": "ratio_referencia_no_valido"})
            continue

        umbral_ratio = fraccion_quiebre * ratio_referencia
        quiebre_tiempo = None
        for i in range(len(tramo)):
            t0 = tramo["timestamp"].iloc[i]
            t_fin = t0 + pd.Timedelta(hours=horas_min_quiebre_sostenido)
            sub_idx = tramo[(tramo["timestamp"] >= t0) & (tramo["timestamp"] <= t_fin)].index
            sub_ratio = ratio.loc[sub_idx].dropna()
            if len(sub_ratio) >= 2 and sub_ratio.max() <= umbral_ratio:
                quiebre_tiempo = t0
                break

        if quiebre_tiempo is None:
            resultados.append({"inicio_riego": e.inicio_riego, "valido": False,
                                "motivo_descarte": "sin_quiebre_relativo_a_et"})
            continue

        meseta = tramo[tramo["timestamp"] >= quiebre_tiempo]
        pmp_evento = meseta["valor"].median()
        resultados.append({
            "inicio_riego": e.inicio_riego,
            "valido": True,
            "quiebre_tiempo": quiebre_tiempo,
            "pmp_evento": round(float(pmp_evento), 2),
            "ratio_referencia": round(float(ratio_referencia), 3),
        })

    return resultados



# ----------------------------------------------------------------------
# Exportacion a Excel
# ----------------------------------------------------------------------
def guardar_excel(df_meteo: pd.DataFrame, ruta_excel: str,
                   incluir_resumen_diario: bool = True,
                   p_tabla: Optional[float] = None, kc: float = 1.0) -> Path:
    """
    Guarda la serie horaria de ET0/precipitacion en un Excel (hoja
    'horario'), con una hoja adicional 'resumen_diario' (ET0 y
    precipitacion diarias, y p_ajustado FAO-56 si se indica `p_tabla`).
    """
    ruta_excel = Path(ruta_excel)
    ruta_excel.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
        df_meteo.to_excel(writer, sheet_name="horario", index=False)

        if incluir_resumen_diario:
            if p_tabla is not None:
                diario = tabla_p_ajustado_diario(df_meteo, p_tabla, kc=kc)
            else:
                diario = (
                    df_meteo.set_index("timestamp")
                    .resample("D")
                    .agg(et0_diaria_mm=("et0", "sum"), precipitacion_diaria_mm=("precipitacion", "sum"))
                    .reset_index()
                )
            diario.to_excel(writer, sheet_name="resumen_diario", index=False)

    _formatear_excel(ruta_excel)
    return ruta_excel


def _formatear_excel(ruta_excel: Path) -> None:
    """Fuente profesional (Arial), cabecera en negrita, columnas autoajustadas y fila de cabecera fija."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    wb = load_workbook(ruta_excel)
    for hoja in wb.worksheets:
        for fila in hoja.iter_rows():
            for celda in fila:
                celda.font = Font(name="Arial", bold=(celda.row == 1))
        for columna in hoja.columns:
            longitud = max((len(str(c.value)) for c in columna if c.value is not None), default=10)
            letra = columna[0].column_letter
            hoja.column_dimensions[letra].width = min(longitud + 2, 40)
        hoja.freeze_panes = "A2"
    wb.save(ruta_excel)


# ----------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------

# ===== KC FENOLOGICO =====


def validar_cultivo(cultivo: dict[str, Any]) -> None:
    from services.locale_numbers import parse_locale_float, parse_p_tabla

    nombre = (cultivo.get("nombre") or "").strip()
    if not nombre:
        raise ValueError("El nombre del cultivo es obligatorio.")

    L1 = parse_locale_float(cultivo.get("L1"))
    L2 = parse_locale_float(cultivo.get("L2"))
    L3 = parse_locale_float(cultivo.get("L3"))
    L4 = parse_locale_float(cultivo.get("L4"))
    kc_ini = parse_locale_float(cultivo.get("kc_ini"))
    kc_med = parse_locale_float(cultivo.get("kc_med"))
    kc_fin = parse_locale_float(cultivo.get("kc_fin"))
    if None in (L1, L2, L3, L4, kc_ini, kc_med, kc_fin):
        raise ValueError("L1–L4 y kc_ini/kc_med/kc_fin deben ser números.")

    if not (0 <= L1 <= L2 <= L3 <= L4):
        raise ValueError("Debe cumplirse 0 ≤ L1 ≤ L2 ≤ L3 ≤ L4.")
    if L4 <= 0:
        raise ValueError("L4 debe ser mayor que 0.")
    if min(kc_ini, kc_med, kc_fin) < 0:
        raise ValueError("Los valores de Kc no pueden ser negativos.")

    p_raw = cultivo.get("p_tabla", None)
    if p_raw is not None and str(p_raw).strip() != "":
        p_tabla = parse_p_tabla(p_raw)
        if p_tabla is None:
            raise ValueError("p_tabla debe ser un número entre 0 y 1.")


def kc_en_dia(
    d: float,
    L1: float,
    L2: float,
    L3: float,
    L4: float,
    kc_ini: float,
    kc_med: float,
    kc_fin: float,
) -> float:
    """Kc FAO-56 por tramos en el día d del ciclo."""
    if d < 0:
        return float(kc_ini)
    if d <= L1:
        return float(kc_ini)
    if d <= L2:
        if L2 == L1:
            return float(kc_med)
        return float(kc_ini + (kc_med - kc_ini) * (d - L1) / (L2 - L1))
    if d <= L3:
        return float(kc_med)
    if d <= L4:
        if L4 == L3:
            return float(kc_fin)
        return float(kc_med + (kc_fin - kc_med) * (d - L3) / (L4 - L3))
    return float(kc_fin)


def etapa_en_dia(d: float, L1: float, L2: float, L3: float, L4: float) -> str:
    if d < 0:
        return "antes de siembra"
    if d <= L1:
        return "inicial"
    if d <= L2:
        return "desarrollo"
    if d <= L3:
        return "media"
    if d <= L4:
        return "final"
    return "post-cosecha"


def dia_ciclo(
    fecha_ref: date,
    fecha_siembra: Optional[date],
    fecha_cosecha: Optional[date],
    L4: float,
    peso_cosecha: float,
) -> Optional[float]:
    """
    Día del ciclo según siembra y/o cosecha.

    peso_cosecha en [0, 1]: 0 = solo siembra, 1 = solo cosecha,
    valores intermedios mezclan ambas estimaciones.
    """
    w = max(0.0, min(1.0, float(peso_cosecha)))
    d_siembra = (fecha_ref - fecha_siembra).days if fecha_siembra is not None else None
    d_cosecha = (L4 - (fecha_cosecha - fecha_ref).days) if fecha_cosecha is not None else None

    if d_siembra is None and d_cosecha is None:
        return None
    if d_siembra is None:
        return float(d_cosecha)  # type: ignore[arg-type]
    if d_cosecha is None:
        return float(d_siembra)
    return float((1.0 - w) * d_siembra + w * d_cosecha)


def puntos_curva_kc(cultivo: dict[str, Any]) -> list[tuple[float, float]]:
    """Vértices de la polilínea FAO-56."""
    return [
        (0.0, float(cultivo["kc_ini"])),
        (float(cultivo["L1"]), float(cultivo["kc_ini"])),
        (float(cultivo["L2"]), float(cultivo["kc_med"])),
        (float(cultivo["L3"]), float(cultivo["kc_med"])),
        (float(cultivo["L4"]), float(cultivo["kc_fin"])),
    ]


def calcular_kc_cultivo(
    cultivo: dict[str, Any],
    fecha_ref: date,
    fecha_siembra: Optional[date],
    fecha_cosecha: Optional[date],
    peso_cosecha: float,
) -> dict[str, Any]:
    d = dia_ciclo(fecha_ref, fecha_siembra, fecha_cosecha, float(cultivo["L4"]), peso_cosecha)
    if d is None:
        return {"kc": None, "dia_ciclo": None, "etapa": None}

    kc = kc_en_dia(
        d,
        float(cultivo["L1"]),
        float(cultivo["L2"]),
        float(cultivo["L3"]),
        float(cultivo["L4"]),
        float(cultivo["kc_ini"]),
        float(cultivo["kc_med"]),
        float(cultivo["kc_fin"]),
    )
    etapa = etapa_en_dia(d, float(cultivo["L1"]), float(cultivo["L2"]), float(cultivo["L3"]), float(cultivo["L4"]))
    return {"kc": round(kc, 4), "dia_ciclo": round(d, 2), "etapa": etapa}


def dibujar_curva_kc(cultivo: dict[str, Any], dia_ciclo_val: Optional[float] = None) -> Figure:
    """Diagrama FAO-56: Kc vs días, marcas L1–L4 con valores numéricos visibles."""
    L1 = float(cultivo["L1"])
    L2 = float(cultivo["L2"])
    L3 = float(cultivo["L3"])
    L4 = float(cultivo["L4"])
    kc_ini = float(cultivo["kc_ini"])
    kc_med = float(cultivo["kc_med"])
    kc_fin = float(cultivo["kc_fin"])

    puntos = puntos_curva_kc(cultivo)
    xs = [p[0] for p in puntos]
    ys = [p[1] for p in puntos]

    fig, ax = plt.subplots(figsize=(8.0, 4.2))
    ax.plot(xs, ys, color="#1a1a1a", linewidth=2.4, solid_capstyle="round")

    for L, label in ((L1, f"L1={L1:g}"), (L2, f"L2={L2:g}"), (L3, f"L3={L3:g}"), (L4, f"L4={L4:g}")):
        ax.axvline(L, color="#c44", linestyle="--", linewidth=1.1, alpha=0.9)
        ax.text(
            L, -0.07, label, color="#c44", ha="center", va="top",
            fontsize=10, fontweight="bold", transform=ax.get_xaxis_transform(),
        )

    y_fase = -0.18
    fases = [
        ((0 + L1) / 2, "inicial"),
        ((L1 + L2) / 2, "desarrollo"),
        ((L2 + L3) / 2, "media"),
        ((L3 + L4) / 2, "final"),
    ]
    for x, texto in fases:
        ax.text(x, y_fase, texto, ha="center", va="top", fontsize=8, color="#555555", transform=ax.get_xaxis_transform())

    ax.annotate(
        f"Kc ini = {kc_ini:.2f}", xy=(L1 / 2, kc_ini), xytext=(0, 10),
        textcoords="offset points", ha="center", fontsize=10, fontweight="bold",
    )
    ax.annotate(
        f"Kc med = {kc_med:.2f}", xy=((L2 + L3) / 2, kc_med), xytext=(0, 10),
        textcoords="offset points", ha="center", fontsize=10, fontweight="bold",
    )
    ax.annotate(
        f"Kc fin = {kc_fin:.2f}", xy=(L4, kc_fin), xytext=(-8, 10),
        textcoords="offset points", ha="right", fontsize=10, fontweight="bold",
    )

    resumen = (
        f"L1={L1:g}  L2={L2:g}  L3={L3:g}  L4={L4:g}\n"
        f"Kc ini={kc_ini:.2f}  Kc med={kc_med:.2f}  Kc fin={kc_fin:.2f}"
    )
    ax.text(
        0.02, 0.98, resumen, transform=ax.transAxes, va="top", ha="left",
        fontsize=9, family="monospace",
        bbox=dict(boxstyle="round,pad=0.4", facecolor="#f5f5f5", edgecolor="#bbbbbb", alpha=0.95),
    )

    if dia_ciclo_val is not None:
        kc_d = kc_en_dia(dia_ciclo_val, L1, L2, L3, L4, kc_ini, kc_med, kc_fin)
        ax.axvline(dia_ciclo_val, color="#2a6fdb", linestyle=":", linewidth=1.2, alpha=0.9)
        ax.scatter(
            [dia_ciclo_val], [kc_d], color="#2a6fdb", s=45, zorder=5,
            label=f"Ref. (d={dia_ciclo_val:.0f}, Kc={kc_d:.2f})",
        )
        ax.legend(loc="lower right", fontsize=8, frameon=False)

    ymax = max(kc_ini, kc_med, kc_fin, 1.0) + 0.30
    ax.set_xlim(0, L4 * 1.05)
    ax.set_ylim(0, max(1.4, ymax))
    ax.set_xlabel("Tiempo (días)")
    ax.set_ylabel("Kc")
    ax.set_title(f"Curva Kc FAO-56 — {cultivo.get('nombre', '')}")
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    fig.subplots_adjust(bottom=0.22, top=0.90)
    return fig


# ===== PIPELINE COMPLETO =====

@dataclass
class InformeCompleto:
    cc_optima: Optional[float]
    n_eventos_detectados: int
    n_eventos_validos: int
    robusto_cc: bool

    pmp_teorico: float
    ad_teorico: Optional[float]

    pmp_operativo: Optional[float]
    fuente_pmp_operativo: str            # "et_normalizado" | "pendiente_vwc" | "ninguno"
    n_candidatos_pmp_et: int
    n_candidatos_pmp_vwc: int

    p_tabla: float
    p_ajustado_medio: Optional[float]
    p_ajustado_min: Optional[float]
    p_ajustado_max: Optional[float]

    raw_p_fijo: Optional[float]
    raw_p_dinamico: Optional[float]
    raw_con_pmp_operativo: Optional[float]

    valle_real_percentil: Optional[float]
    umbral_inferior_recomendado: Optional[float]
    fuente_umbral_inferior: Optional[str]

    coef_seguridad_vwc: float = 2.0
    umbral_superior_final: Optional[float] = None

    diagnostico_lluvia: pd.DataFrame = field(default_factory=pd.DataFrame)
    eventos_cc: list = field(default_factory=list)
    eventos_pmp_et: list = field(default_factory=list)
    df_meteo: Optional[pd.DataFrame] = None
    tabla_p: Optional[pd.DataFrame] = None
    avisos: list[str] = field(default_factory=list)
    sensor_fecha_inicio: Optional[pd.Timestamp] = None
    sensor_fecha_fin: Optional[pd.Timestamp] = None
    meteo_fecha_inicio: Optional[pd.Timestamp] = None
    meteo_fecha_fin: Optional[pd.Timestamp] = None
    fuente_meteo: str = "open_meteo_api"  # "csv" | "open_meteo_api"


# ----------------------------------------------------------------------
# Utilidades internas
# ----------------------------------------------------------------------
def _resolver_pmp_teorico(textura: Optional[str], pmp_teorico: Optional[float]) -> float:
    if pmp_teorico is not None:
        return pmp_teorico
    if textura is None:
        raise ValueError("Indica 'textura' o 'pmp_teorico'.")
    clave = textura.strip().lower()
    if clave not in TABLA_TEXTURAS:
        raise ValueError(f"Textura '{textura}' no reconocida. Opciones: {list(TABLA_TEXTURAS)}")
    return TABLA_TEXTURAS[clave]["pmp_teorica"]


def _elegir_modo_meteo(fecha_inicio: pd.Timestamp, fecha_fin: pd.Timestamp):
    """Decide si usar el endpoint de forecast (past_days) o el historico (archive), segun antigüedad."""
    hoy = date.today()
    dias_atras = (hoy - fecha_inicio.date()).days
    if dias_atras <= 90:
        past_days = max(dias_atras + 2, 1)  # margen de 2 dias
        forecast_days = max((fecha_fin.date() - hoy).days + 1, 1)
        return "forecast", past_days, forecast_days
    return "historico", None, None


def _valores_previos_a_cada_riego(df: pd.DataFrame, inicios: list[pd.Timestamp]) -> list[float]:
    """Para cada riego (salvo el primero), el ultimo valor registrado justo antes de que empiece."""
    inicios_ordenados = sorted(inicios)
    valores = []
    for t in inicios_ordenados[1:]:
        previos = df[df["timestamp"] < t]
        if not previos.empty:
            valores.append(float(previos["valor"].iloc[-1]))
    return valores


# ----------------------------------------------------------------------
# Pipeline principal
# ----------------------------------------------------------------------
def ejecutar_analisis_completo(
    csv_path: str,
    p_tabla: float,
    lat: float,
    lon: float,
    fecha_inicio: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    textura: Optional[str] = None,
    pmp_teorico: Optional[float] = None,
    kc: float = 1.0,
    con_cabecera: bool = False,
    inicios_riego_manual: Optional[list[str]] = None,
    umbral_lluvia_mm: float = 1.0,
    excluir_posible_lluvia: bool = False,
    percentil_valle: float = 75.0,
    coef_seguridad_vwc: float = 2.0,
    params_cc: Optional[ParametrosDeteccion] = None,
    params_pmp: Optional[ParametrosPMP] = None,
    col_timestamp: int = 5,
    col_valor: int = 6,
    df_meteo: Optional[pd.DataFrame] = None,
    _obtener_meteo_fn=obtener_et_precipitacion,   # inyectable para pruebas sin red
) -> InformeCompleto:

    pmp_teorico_val = _resolver_pmp_teorico(textura, pmp_teorico)
    params_cc = params_cc or ParametrosDeteccion()
    params_pmp = params_pmp or ParametrosPMP()
    avisos: list[str] = []

    # 1. Carga y recorte de la serie del sensor
    df = cargar_serie(csv_path, col_timestamp=col_timestamp, col_valor=col_valor, con_cabecera=con_cabecera)
    if fecha_inicio:
        df = df[df["timestamp"] >= pd.to_datetime(fecha_inicio)]
    if fecha_fin:
        df = df[df["timestamp"] <= pd.to_datetime(fecha_fin) + pd.Timedelta(hours=23, minutes=59)]
    df = df.reset_index(drop=True)
    if df.empty:
        raise ValueError("No quedan datos del sensor tras aplicar el rango de fechas indicado.")

    fecha_inicio_ts = pd.to_datetime(fecha_inicio) if fecha_inicio else df["timestamp"].min()
    fecha_fin_ts = pd.to_datetime(fecha_fin) if fecha_fin else df["timestamp"].max()

    # 2. Deteccion de riegos y CC optima
    if inicios_riego_manual:
        inicios = [pd.to_datetime(t) for t in inicios_riego_manual]
    else:
        inicios = detectar_inicios_riego_automatico(df, params_cc)

    eventos = analizar_eventos(df, inicios, params_cc)
    eventos_validos = [e for e in eventos if e.valido]
    cc_optima = float(np.median([e.cc_evento for e in eventos_validos])) if eventos_validos else None
    robusto_cc = len(eventos_validos) >= params_cc.n_min_eventos

    # 3. ET0 y precipitacion (CSV local o Open-Meteo API)
    fuente_meteo = "csv" if df_meteo is not None else "open_meteo_api"
    if df_meteo is not None:
        df_meteo_src = df_meteo.copy()
        if "timestamp" not in df_meteo_src.columns:
            raise ValueError("df_meteo debe tener columnas timestamp, et0, precipitacion.")
        df_meteo_src["timestamp"] = _ensure_datetime64_ns(df_meteo_src["timestamp"])
        aviso = aviso_solape_meteo_sensor(df["timestamp"].min(), df["timestamp"].max(), df_meteo_src)
        if aviso:
            avisos.append(aviso)
    else:
        modo, past_days, forecast_days = _elegir_modo_meteo(fecha_inicio_ts, fecha_fin_ts)
        if modo == "forecast":
            df_meteo_src = _obtener_meteo_fn(lat, lon, past_days=past_days, forecast_days=forecast_days)
        else:
            df_meteo_src = _obtener_meteo_fn(
                lat, lon, usar_archivo_historico=True,
                fecha_inicio=fecha_inicio_ts.strftime("%Y-%m-%d"),
                fecha_fin=fecha_fin_ts.strftime("%Y-%m-%d"),
            )
        df_meteo_src["timestamp"] = _ensure_datetime64_ns(df_meteo_src["timestamp"])

    df_meteo = df_meteo_src[
        (df_meteo_src["timestamp"] >= fecha_inicio_ts)
        & (df_meteo_src["timestamp"] <= fecha_fin_ts + pd.Timedelta(hours=23))
    ].reset_index(drop=True)
    if df_meteo.empty:
        raise ValueError(
            "El CSV meteo no cubre el periodo del sensor (solape nulo). "
            "Descarga un rango Open-Meteo que solape las fechas del sensor."
        )

    # 4. Riego vs lluvia
    diagnostico_lluvia = marcar_posible_lluvia(eventos, df_meteo, umbral_mm=umbral_lluvia_mm)

    if excluir_posible_lluvia and not diagnostico_lluvia.empty:
        inicios_lluvia = set(diagnostico_lluvia.loc[diagnostico_lluvia["posible_lluvia"], "inicio_riego"])
        eventos = [e for e in eventos if e.inicio_riego not in inicios_lluvia]
        eventos_validos = [e for e in eventos if e.valido]
        if eventos_validos:
            cc_optima = float(np.median([e.cc_evento for e in eventos_validos]))
            robusto_cc = len(eventos_validos) >= params_cc.n_min_eventos
        else:
            cc_optima = None
            robusto_cc = False

    # 5. PMP operativo (dos criterios)
    resultado_pmp_vwc = analizar_pmp_operativo(df, eventos, params_pmp)
    resultados_pmp_et = analizar_pmp_operativo_normalizado_et(df, df_meteo, eventos, kc=kc)
    validos_et = [r for r in resultados_pmp_et if r.get("valido")]

    if validos_et:
        pmp_operativo = float(np.median([r["pmp_evento"] for r in validos_et]))
        fuente_pmp_operativo = "et_normalizado"
    elif resultado_pmp_vwc.pmp_operativo is not None:
        pmp_operativo = resultado_pmp_vwc.pmp_operativo
        fuente_pmp_operativo = "pendiente_vwc (menos fiable: verificar resolucion del sensor)"
    else:
        pmp_operativo = None
        fuente_pmp_operativo = "ninguno"

    # 6. Ajuste dinamico de p
    tabla_p = tabla_p_ajustado_diario(df_meteo, p_tabla, kc=kc) if not df_meteo.empty else None
    if tabla_p is not None and not tabla_p.empty:
        p_medio = float(tabla_p["p_ajustado"].mean())
        p_min = float(tabla_p["p_ajustado"].min())
        p_max = float(tabla_p["p_ajustado"].max())
    else:
        p_medio = p_min = p_max = None

    # AD y RAW
    ad_teorico = (cc_optima - pmp_teorico_val) if cc_optima is not None else None
    raw_p_fijo = (cc_optima - p_tabla * ad_teorico) if ad_teorico is not None else None
    raw_p_dinamico = (cc_optima - p_medio * ad_teorico) if (ad_teorico is not None and p_medio is not None) else None

    raw_con_pmp_operativo = None
    if cc_optima is not None and pmp_operativo is not None:
        ad_operativo = cc_optima - pmp_operativo
        raw_con_pmp_operativo = cc_optima - p_tabla * ad_operativo

    # 7. Umbral inferior = max entre RAWs disponibles (el valle real NO entra)
    valores_previos = _valores_previos_a_cada_riego(df, inicios)
    valle_real_percentil = float(np.percentile(valores_previos, percentil_valle)) if valores_previos else None

    # El RAW con PMP operativo solo entra en el max si el candidato viene del
    # criterio normalizado por ET (el fiable). Si solo lo encontro el
    # criterio de pendiente absoluta de VWC, no se usa para el umbral
    # -- se queda solo como dato informativo en el informe.
    if fuente_pmp_operativo == "et_normalizado":
        candidatos_raw = [
            ("RAW con PMP operativo", raw_con_pmp_operativo),
            ("RAW con p dinamico (media)", raw_p_dinamico),
            ("RAW con p fijo", raw_p_fijo),
        ]
    else:
        candidatos_raw = [
            ("RAW con p dinamico (media)", raw_p_dinamico),
            ("RAW con p fijo", raw_p_fijo),
        ]
    candidatos_raw = [(nombre, v) for nombre, v in candidatos_raw if v is not None]

    if candidatos_raw:
        fuente_umbral_inferior, umbral_inferior_recomendado = max(candidatos_raw, key=lambda x: x[1])
    else:
        fuente_umbral_inferior, umbral_inferior_recomendado = None, None

    umbral_superior_final = (cc_optima + coef_seguridad_vwc) if cc_optima is not None else None

    sensor_fecha_inicio = pd.Timestamp(df["timestamp"].min())
    sensor_fecha_fin = pd.Timestamp(df["timestamp"].max())
    meteo_fecha_inicio = pd.Timestamp(df_meteo["timestamp"].min()) if not df_meteo.empty else None
    meteo_fecha_fin = pd.Timestamp(df_meteo["timestamp"].max()) if not df_meteo.empty else None

    return InformeCompleto(
        cc_optima=cc_optima,
        n_eventos_detectados=len(eventos),
        n_eventos_validos=len(eventos_validos),
        robusto_cc=robusto_cc,
        pmp_teorico=pmp_teorico_val,
        ad_teorico=ad_teorico,
        pmp_operativo=pmp_operativo,
        fuente_pmp_operativo=fuente_pmp_operativo,
        n_candidatos_pmp_et=len(validos_et),
        n_candidatos_pmp_vwc=resultado_pmp_vwc.n_eventos_validos,
        p_tabla=p_tabla,
        p_ajustado_medio=p_medio,
        p_ajustado_min=p_min,
        p_ajustado_max=p_max,
        raw_p_fijo=raw_p_fijo,
        raw_p_dinamico=raw_p_dinamico,
        raw_con_pmp_operativo=raw_con_pmp_operativo,
        valle_real_percentil=valle_real_percentil,
        umbral_inferior_recomendado=umbral_inferior_recomendado,
        fuente_umbral_inferior=fuente_umbral_inferior,
        coef_seguridad_vwc=coef_seguridad_vwc,
        umbral_superior_final=umbral_superior_final,
        diagnostico_lluvia=diagnostico_lluvia,
        eventos_cc=eventos,
        eventos_pmp_et=resultados_pmp_et,
        df_meteo=df_meteo,
        tabla_p=tabla_p,
        avisos=avisos,
        sensor_fecha_inicio=sensor_fecha_inicio,
        sensor_fecha_fin=sensor_fecha_fin,
        meteo_fecha_inicio=meteo_fecha_inicio,
        meteo_fecha_fin=meteo_fecha_fin,
        fuente_meteo=fuente_meteo,
    )


# ----------------------------------------------------------------------
# Informe en consola
# ----------------------------------------------------------------------
def _fmt_informe_ts(ts: Optional[pd.Timestamp]) -> str:
    if ts is None or pd.isna(ts):
        return "n/d"
    t = pd.Timestamp(ts)
    return t.strftime("%d/%m/%Y %H:%M")


def imprimir_informe_completo(inf: InformeCompleto) -> None:
    print("=" * 70)
    print("INFORME COMPLETO DE UMBRALES DE RIEGO")
    print("=" * 70)

    print("\n[DATOS CONSIDERADOS]")
    print(
        f"  Sensor (humedad) : {_fmt_informe_ts(inf.sensor_fecha_inicio)} → "
        f"{_fmt_informe_ts(inf.sensor_fecha_fin)}"
    )
    fuente_label = "CSV Open-Meteo" if inf.fuente_meteo == "csv" else "API Open-Meteo"
    if inf.meteo_fecha_inicio is not None and inf.meteo_fecha_fin is not None:
        print(
            f"  Meteo (ET0/lluvia): {_fmt_informe_ts(inf.meteo_fecha_inicio)} → "
            f"{_fmt_informe_ts(inf.meteo_fecha_fin)}  (fuente: {fuente_label})"
        )
    else:
        print(f"  Meteo (ET0/lluvia): n/d  (fuente: {fuente_label})")

    if inf.avisos:
        print("\n[AVISOS]")
        for aviso in inf.avisos:
            print(f"  - {aviso}")

    print(f"\n[CC OPTIMA / UMBRAL SUPERIOR]")
    print(f"  Eventos detectados/validos : {inf.n_eventos_detectados} / {inf.n_eventos_validos}"
          f" ({'robusto' if inf.robusto_cc else 'NO robusto, ampliar periodo'})")
    print(f"  CC optima                  : {inf.cc_optima:.2f} %VWC" if inf.cc_optima is not None else "  CC optima: sin datos")

    print(f"\n[RIEGO vs LLUVIA]")
    if not inf.diagnostico_lluvia.empty:
        n_lluvia = int(inf.diagnostico_lluvia["posible_lluvia"].sum())
        print(f"  {n_lluvia} de {len(inf.diagnostico_lluvia)} eventos podrian ser lluvia (>= umbral de precipitacion)")
        print(inf.diagnostico_lluvia.to_string(index=False))
    else:
        print("  Sin datos meteorologicos para contrastar.")

    print(f"\n[PMP]")
    print(f"  PMP teorico (tabla/textura): {inf.pmp_teorico:.2f} %VWC")
    print(f"  PMP operativo (proxy)      : "
          f"{f'{inf.pmp_operativo:.2f} %VWC' if inf.pmp_operativo is not None else 'no encontrado en este periodo'}")
    print(f"  Fuente del PMP operativo   : {inf.fuente_pmp_operativo}")
    print(f"  Candidatos (ET-normalizado): {inf.n_candidatos_pmp_et}  |  Candidatos (pendiente VWC): {inf.n_candidatos_pmp_vwc}")

    print(f"\n[AJUSTE DINAMICO DE p (FAO-56)]")
    print(f"  p de tabla (fijo)  : {inf.p_tabla:.3f}")
    if inf.p_ajustado_medio is not None:
        print(f"  p_ajustado del periodo: media={inf.p_ajustado_medio:.3f} "
              f"(min={inf.p_ajustado_min:.3f}, max={inf.p_ajustado_max:.3f})")

    print(f"\n[UMBRAL INFERIOR (RAW) - distintas versiones]")
    print(f"  AD teorico (CC_optima - PMP_teorico): "
          f"{f'{inf.ad_teorico:.2f} %VWC' if inf.ad_teorico is not None else 'n/d'}")
    print(f"  RAW con p de tabla fijo    : {f'{inf.raw_p_fijo:.2f} %VWC' if inf.raw_p_fijo is not None else 'n/d'}")
    print(f"  RAW con p_ajustado (media) : {f'{inf.raw_p_dinamico:.2f} %VWC' if inf.raw_p_dinamico is not None else 'n/d'}")
    print(f"  RAW con PMP operativo      : "
          f"{f'{inf.raw_con_pmp_operativo:.2f} %VWC' if inf.raw_con_pmp_operativo is not None else 'n/d (sin PMP operativo robusto)'}")

    print(f"\n[UMBRAL INFERIOR OPERATIVO RECOMENDADO]")
    print(f"  Valle real (percentil de seguridad): "
          f"{f'{inf.valle_real_percentil:.2f} %VWC' if inf.valle_real_percentil is not None else 'n/d'}")
    print(f"  -> RECOMENDADO (max entre RAWs; valle real no entra): "
          f"{f'{inf.umbral_inferior_recomendado:.2f} %VWC' if inf.umbral_inferior_recomendado is not None else 'n/d'}")
    if inf.fuente_umbral_inferior:
        print(f"     (elegido: {inf.fuente_umbral_inferior})")

    print(f"\n[RECOMENDACION]")
    if inf.umbral_superior_final is not None and inf.cc_optima is not None:
        print(f"  Umbral superior final : {inf.umbral_superior_final:.2f} %VWC")
        print(f"    (CC optima {inf.cc_optima:.2f} + factor seguridad {inf.coef_seguridad_vwc:.2f} %VWC)")
    else:
        print("  Umbral superior final : n/d")

    if inf.umbral_inferior_recomendado is not None:
        print(f"  Umbral inferior       : {inf.umbral_inferior_recomendado:.2f} %VWC")
        if inf.fuente_umbral_inferior:
            print(f"    (elegido: {inf.fuente_umbral_inferior} — max entre RAWs disponibles; valle real no entra)")
    else:
        print("  Umbral inferior       : n/d")

    print(f"  Valle real            : "
          f"{f'{inf.valle_real_percentil:.2f} %VWC' if inf.valle_real_percentil is not None else 'n/d'}")

    print("\n  Nota comparativa:")
    if inf.valle_real_percentil is not None and inf.umbral_inferior_recomendado is not None:
        diff = inf.valle_real_percentil - inf.umbral_inferior_recomendado
        if diff >= 1.0:
            print("    valle real ≫ recomendado → el histórico de riego es más conservador que el modelo;")
            print("    hay margen para espaciar riegos si se quiere optimizar consumo de agua.")
        elif diff <= -1.0:
            print("    valle real ≪ recomendado → ATENCIÓN: el histórico muestra riegos iniciados")
            print("    por debajo del umbral que el modelo considera seguro; revisar si ha habido")
            print("    estrés hídrico no detectado.")
        else:
            print("    valle real y umbral recomendado están alineados (diferencia < 1 %VWC).")
    else:
        print("    No hay suficientes datos para comparar valle real y umbral recomendado.")

    print("=" * 70)


# ----------------------------------------------------------------------
# Exportacion a Excel consolidado
# ----------------------------------------------------------------------
def guardar_excel_completo(inf: InformeCompleto, ruta_excel: str) -> Path:
    ruta_excel = Path(ruta_excel)
    ruta_excel.parent.mkdir(parents=True, exist_ok=True)

    resumen = pd.DataFrame([
        {"campo": "CC optima %VWC", "valor": inf.cc_optima},
        {"campo": "Coeficiente de seguridad %VWC", "valor": inf.coef_seguridad_vwc},
        {"campo": "Umbral superior final %VWC", "valor": inf.umbral_superior_final},
        {"campo": "Eventos detectados/validos", "valor": f"{inf.n_eventos_detectados}/{inf.n_eventos_validos}"},
        {"campo": "Robusto (CC)", "valor": inf.robusto_cc},
        {"campo": "PMP teorico %VWC", "valor": inf.pmp_teorico},
        {"campo": "PMP operativo (proxy) %VWC", "valor": inf.pmp_operativo},
        {"campo": "Fuente PMP operativo", "valor": inf.fuente_pmp_operativo},
        {"campo": "p de tabla", "valor": inf.p_tabla},
        {"campo": "p_ajustado medio del periodo", "valor": inf.p_ajustado_medio},
        {"campo": "AD teorico %VWC", "valor": inf.ad_teorico},
        {"campo": "RAW con p fijo %VWC", "valor": inf.raw_p_fijo},
        {"campo": "RAW con p dinamico (media) %VWC", "valor": inf.raw_p_dinamico},
        {"campo": "RAW con PMP operativo %VWC", "valor": inf.raw_con_pmp_operativo},
        {"campo": "Valle real (percentil seguridad) %VWC", "valor": inf.valle_real_percentil},
        {"campo": "UMBRAL INFERIOR RECOMENDADO %VWC", "valor": inf.umbral_inferior_recomendado},
        {"campo": "Fuente umbral inferior", "valor": inf.fuente_umbral_inferior},
    ])

    eventos_cc_df = pd.DataFrame([{
        "inicio_riego": e.inicio_riego, "pico_valor": e.pico_valor, "pico_tiempo": e.pico_tiempo,
        "estabilizacion_tiempo": e.estabilizacion_tiempo, "cc_evento": e.cc_evento,
        "valido": e.valido, "motivo_descarte": e.motivo_descarte,
    } for e in inf.eventos_cc])

    pmp_et_df = pd.DataFrame(inf.eventos_pmp_et) if inf.eventos_pmp_et else pd.DataFrame()

    with pd.ExcelWriter(ruta_excel, engine="openpyxl") as writer:
        resumen.to_excel(writer, sheet_name="resumen", index=False)
        eventos_cc_df.to_excel(writer, sheet_name="eventos_cc", index=False)
        inf.diagnostico_lluvia.to_excel(writer, sheet_name="riego_vs_lluvia", index=False)
        pmp_et_df.to_excel(writer, sheet_name="pmp_operativo_et", index=False)
        if inf.df_meteo is not None:
            inf.df_meteo.to_excel(writer, sheet_name="meteo_horario", index=False)
        if inf.tabla_p is not None:
            inf.tabla_p.to_excel(writer, sheet_name="meteo_diario_p_ajustado", index=False)

    _formatear_excel(ruta_excel)
    return ruta_excel


def guardar_excel_completo_bytes(informe: InformeCompleto) -> bytes:
    """Genera el Excel del informe en memoria (para descarga Streamlit)."""
    import tempfile

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        ruta = tmp.name
    try:
        guardar_excel_completo(informe, ruta)
        return Path(ruta).read_bytes()
    finally:
        with contextlib.suppress(Exception):
            Path(ruta).unlink(missing_ok=True)
