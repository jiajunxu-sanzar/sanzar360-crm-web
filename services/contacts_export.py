"""Export contact sensor overview table to Excel and PDF."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pandas as pd

_EXPORT_HEADERS = [
    "Contacto",
    "Sensores",
    "Último contacto",
    "Próxima acción",
    "Incidencias",
]
EXPORT_HEADERS = _EXPORT_HEADERS
SHEET_NAME = "Contactos con sensores"
TITLE = "Contactos — Sensores e incidencias"
HEADER_FILL = "2D6A4F"
HEADER_FONT = "FFFFFF"
ROW_FILL_VERDE = "C6F6D5"
ROW_FILL_AMARILLO = "FEF3C7"
MAX_COL_WIDTH = 60


def _export_filename(ext: str, exported_at: datetime) -> str:
    return f"contactos_sensores_{exported_at.strftime('%Y%m%d')}.{ext}"


def _format_ultimo_contacto_export(fecha: object, canal: object, detalle: object = "") -> str:
    detail = str(detalle or "").strip()
    if detail:
        return detail
    f = str(fecha or "").strip()
    c = str(canal or "").strip()
    if f and c:
        return f"{f} ({c})"
    return f or c or "—"


def _format_proxima_accion_export(fecha: object, detalle: object) -> str:
    f = str(fecha or "").strip()
    d = str(detalle or "").strip()
    if f and d:
        return f"{f} · {d}"
    return f or d or "—"


def _format_incidencias_export(count: object, detalle: object = "") -> str:
    text = str(detalle or "").strip()
    if text:
        return text
    try:
        n = int(count)
    except (TypeError, ValueError):
        n = 0
    if n <= 0:
        return "—"
    label = "abierta" if n == 1 else "abiertas"
    return f"{n} {label}"


def _format_sensors_export(sensor_sns: object) -> str:
    text = str(sensor_sns or "").strip()
    return text or "—"


def _semaforo_fill_hex(semaforo: str) -> str | None:
    sem = (semaforo or "").strip().lower()
    if sem == "verde":
        return ROW_FILL_VERDE
    if sem == "amarillo":
        return ROW_FILL_AMARILLO
    return None


def _overview_rows(overview_df: pd.DataFrame) -> list[list[object]]:
    if overview_df is None or overview_df.empty:
        return []
    rows: list[list[object]] = []
    for record in overview_df.fillna("").to_dict("records"):
        rows.append(
            [
                str(record.get("nombre", "") or "Sin nombre"),
                _format_sensors_export(record.get("sensor_sns")),
                _format_ultimo_contacto_export(
                    record.get("ultimo_contacto"),
                    record.get("ultimo_contacto_canal"),
                    record.get("ultimo_contacto_detalle"),
                ),
                _format_proxima_accion_export(
                    record.get("proxima_accion_fecha"),
                    record.get("proxima_accion_detalle"),
                ),
                _format_incidencias_export(
                    record.get("incidencias_abiertas"),
                    record.get("incidencias_detalle"),
                ),
            ]
        )
    return rows


def _semaforo_values(overview_df: pd.DataFrame) -> list[str]:
    if overview_df is None or overview_df.empty:
        return []
    return [str(v or "").strip().lower() for v in overview_df.get("semaforo", pd.Series(dtype=str)).tolist()]


def build_overview_xlsx_bytes(
    overview_df: pd.DataFrame,
    exported_at: datetime | None = None,
) -> tuple[bytes, str]:
    """Return (xlsx_bytes, suggested_filename_without_path)."""
    now = exported_at or datetime.now()
    filename = _export_filename("xlsx", now)

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME

    title_text = f"{TITLE} — Generado el {now.strftime('%d/%m/%Y %H:%M')}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(_EXPORT_HEADERS))
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(fill_type="solid", fgColor=HEADER_FILL)
    header_font = Font(bold=True, color=HEADER_FONT)

    for col_idx, header in enumerate(_EXPORT_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_rows = _overview_rows(overview_df)
    semaforos = _semaforo_values(overview_df)
    for row_offset, (row_values, semaforo) in enumerate(zip(data_rows, semaforos), start=3):
        row_fill_hex = _semaforo_fill_hex(semaforo)
        row_fill = PatternFill(fill_type="solid", fgColor=row_fill_hex) if row_fill_hex else None
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row=row_offset, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if row_fill is not None:
                cell.fill = row_fill

    col_widths: list[int] = []
    for col_idx, header in enumerate(_EXPORT_HEADERS, start=1):
        max_len = len(header)
        for row_offset in range(3, 3 + len(data_rows)):
            value = ws.cell(row=row_offset, column=col_idx).value
            max_len = max(max_len, len(str(value or "")))
        col_widths.append(min(max_len + 2, MAX_COL_WIDTH))
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=2, column=col_idx).column_letter].width = width

    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue(), filename


def _escape_xml(s: str) -> str:
    return (
        (s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def build_overview_pdf_bytes(
    overview_df: pd.DataFrame,
    exported_at: datetime | None = None,
) -> tuple[bytes, str]:
    """Return (pdf_bytes, suggested_filename_without_path)."""
    now = exported_at or datetime.now()
    filename = _export_filename("pdf", now)

    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ModuleNotFoundError:
        minimal = (
            "%PDF-1.4\n1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj\n"
            "2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj\n"
            "3 0 obj<</Type/Page/MediaBox[0 0 612 792]>>endobj\n"
            "trailer<</Root 1 0 R>>\n%%EOF\n"
        )
        return minimal.encode("latin-1"), filename

    buf = BytesIO()
    page_size = landscape(A4)
    doc = SimpleDocTemplate(
        buf,
        pagesize=page_size,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.2 * cm,
        bottomMargin=1.2 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        name="ContactsExportTitle",
        parent=styles["Heading1"],
        fontSize=16,
        spaceAfter=6,
    )
    cell_style = ParagraphStyle(
        name="ContactsExportCell",
        parent=styles["Normal"],
        fontSize=8,
        leading=10,
    )

    story: list = []
    story.append(Paragraph(_escape_xml(TITLE), title_style))
    story.append(
        Paragraph(
            _escape_xml(f"Generado el {now.strftime('%d/%m/%Y %H:%M')}"),
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.4 * cm))

    data_rows = _overview_rows(overview_df)
    semaforos = _semaforo_values(overview_df)
    tbl_data: list[list[object]] = [_EXPORT_HEADERS.copy()]
    for row_values in data_rows:
        tbl_data.append([Paragraph(_escape_xml(str(v)), cell_style) for v in row_values])

    tw = doc.width
    tbl = Table(
        tbl_data,
        colWidths=[
            tw * 0.20,
            tw * 0.28,
            tw * 0.18,
            tw * 0.18,
            tw * 0.16,
        ],
        repeatRows=1,
    )

    style_commands: list = [
        ("FONT", (0, 0), (-1, -1), "Helvetica", 8),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{HEADER_FILL}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    for row_idx, semaforo in enumerate(semaforos, start=1):
        fill_hex = _semaforo_fill_hex(semaforo)
        if fill_hex:
            style_commands.append(
                ("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor(f"#{fill_hex}"))
            )
    tbl.setStyle(TableStyle(style_commands))
    story.append(tbl)

    doc.build(story)
    return buf.getvalue(), filename
