"""Tests for contact overview Excel/PDF export."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

import pandas as pd

from services.contacts_export import (
    EXPORT_HEADERS,
    SHEET_NAME,
    build_overview_pdf_bytes,
    build_overview_xlsx_bytes,
)


def _overview_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "contact_id": "c1",
                "nombre": "Cliente Verde",
                "num_sensores": 2,
                "sensor_sns": "UC001, EM002",
                "ultimo_contacto": "01/06/2025",
                "ultimo_contacto_canal": "llamada",
                "proxima_accion_fecha": "10/06/2025",
                "proxima_accion_detalle": "Llamar",
                "incidencias_abiertas": 1,
                "semaforo": "verde",
            },
            {
                "contact_id": "c2",
                "nombre": "Cliente Amarillo",
                "num_sensores": 1,
                "sensor_sns": "UC003",
                "ultimo_contacto": "02/06/2025",
                "ultimo_contacto_canal": "email",
                "proxima_accion_fecha": "15/06/2025",
                "proxima_accion_detalle": "Visita",
                "incidencias_abiertas": 2,
                "semaforo": "amarillo",
            },
            {
                "contact_id": "c3",
                "nombre": "Sin sensores",
                "num_sensores": 0,
                "sensor_sns": "",
                "ultimo_contacto": "",
                "ultimo_contacto_canal": "",
                "proxima_accion_fecha": "",
                "proxima_accion_detalle": "",
                "incidencias_abiertas": 0,
                "semaforo": "sin_sensores",
            },
        ]
    )


_EXPORTED_AT = datetime(2025, 6, 18, 14, 30)


def test_build_overview_xlsx_bytes_structure() -> None:
    overview = _overview_df()
    content, _fname = build_overview_xlsx_bytes(overview, exported_at=_EXPORTED_AT)

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content))
    assert SHEET_NAME in wb.sheetnames
    ws = wb[SHEET_NAME]
    headers = [ws.cell(row=2, column=col).value for col in range(1, len(EXPORT_HEADERS) + 1)]
    assert headers == EXPORT_HEADERS
    assert ws.max_row == 2 + len(overview)


def test_build_overview_xlsx_bytes_semaforo_fill() -> None:
    overview = _overview_df()
    content, _fname = build_overview_xlsx_bytes(overview, exported_at=_EXPORTED_AT)

    from openpyxl import load_workbook

    ws = load_workbook(BytesIO(content))[SHEET_NAME]
    verde_fill = ws.cell(row=3, column=1).fill.fgColor.rgb
    amarillo_fill = ws.cell(row=4, column=1).fill.fgColor.rgb
    neutral_fill = ws.cell(row=5, column=1).fill
    assert verde_fill.endswith("C6F6D5")
    assert amarillo_fill.endswith("FEF3C7")
    assert neutral_fill.fill_type is None or neutral_fill.fgColor.rgb in ("00000000", "FFFFFFFF", None)


def test_build_overview_pdf_bytes_returns_pdf() -> None:
    overview = _overview_df()
    content, fname = build_overview_pdf_bytes(overview, exported_at=_EXPORTED_AT)
    assert fname.endswith(".pdf")
    assert len(content) > 100
    assert content[:4] == b"%PDF"


def test_export_filenames_use_exported_at() -> None:
    overview = _overview_df()
    _, xlsx_name = build_overview_xlsx_bytes(overview, exported_at=_EXPORTED_AT)
    _, pdf_name = build_overview_pdf_bytes(overview, exported_at=_EXPORTED_AT)
    assert xlsx_name == "contactos_sensores_20250618.xlsx"
    assert pdf_name == "contactos_sensores_20250618.pdf"
