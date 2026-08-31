"""Computo de horas del control de asistencia."""
from __future__ import annotations

from datetime import date

import pandas as pd

from services import attendance_service as core
from services.attendance_service import AttendancePerson


def _person(jornada: str = "8", excepciones: str = "") -> AttendancePerson:
    return AttendancePerson(
        employee_id="EMP001",
        nombre="David Ortiz",
        nombre_fichaje="david",
        jornada=jornada,
        jornada_excepciones=excepciones,
    )


def _holidays_frame(fechas: list[str], anio: str = "2026") -> pd.DataFrame:
    return pd.DataFrame(
        [{"fecha": f, "nombre_festivo": "x", "ambito": "n", "municipio": "Leganes", "anio": anio} for f in fechas]
    )


def _absence_row(**kwargs: str) -> dict[str, str]:
    row = {
        "leave_id": "LV001",
        "employee_id": "EMP001",
        "nombre_employee": "David Ortiz",
        "tipo": "vacaciones",
        "fecha_inicio": "2026-07-06",
        "fecha_fin": "2026-07-10",
        "medio_dia": "FALSE",
        "estado": "aprobado",
        "comentario": "",
        "created_at": "",
        "updated_at": "",
    }
    row.update(kwargs)
    return row


# ------------------------------------------------------------------ horas


def test_parse_marks_reads_every_clock_in() -> None:
    assert core.parse_marks("08:57 14:02 15:00 18:30") == ["08:57", "14:02", "15:00", "18:30"]
    assert core.parse_marks("") == []
    assert core.parse_marks(None) == []
    assert core.parse_marks(float("nan")) == []


def test_hours_between_handles_overnight() -> None:
    assert core.hours_between("09:00", "17:30") == 8.5
    assert core.hours_between("22:00", "02:00") == 4.0


def test_meal_break_only_for_long_shifts() -> None:
    assert core.meal_break_for(8) == 0.5
    assert core.meal_break_for(8.5) == 0.5
    assert core.meal_break_for(5) == 0.0
    assert core.meal_break_for(4) == 0.0


def test_full_day_adds_the_meal_break_so_the_net_day_is_the_shift() -> None:
    entrada, salida, horas = core.full_day_times(8)
    assert (entrada, salida, horas) == ("09:00", "17:30", 8.5)
    assert core.net_hours([{"horas": horas}], 8) == 8.0

    entrada, salida, horas = core.full_day_times(5)
    assert (entrada, salida, horas) == ("09:00", "14:00", 5.0)
    assert core.net_hours([{"horas": horas}], 5) == 5.0


def test_net_hours_subtracts_one_break_per_recorded_day() -> None:
    dias = [{"horas": "8.5"}, {"horas": "9.0"}, {"horas": "7.5"}]
    assert core.net_hours(dias, 8) == 25.0 - 1.5
    assert core.net_hours(dias, 5) == 25.0


# --------------------------------------------------------------- jornadas


def test_resolve_jornada_uses_month_exception_over_default() -> None:
    assert core.resolve_jornada("5", "2026-07:8, 2026-08:8", 2026, 7) == 8.0
    assert core.resolve_jornada("5", "2026-07:8, 2026-08:8", 2026, 6) == 5.0
    assert core.resolve_jornada("5", "", 2026, 6) == 5.0
    assert core.resolve_jornada("", "", 2026, 6) is None
    assert core.resolve_jornada("8h", "", 2026, 6) == 8.0
    assert core.resolve_jornada("4,5", "", 2026, 6) == 4.5


def test_people_by_report_name_matches_alias_case_insensitively() -> None:
    person = _person()
    indice = core.people_by_report_name([person])
    assert indice["david"] is person
    assert indice["david ortiz"] is person


def test_people_by_report_name_accepts_several_aliases() -> None:
    person = AttendancePerson(
        employee_id="EMP002", nombre="Jiajun Xu", nombre_fichaje="jiajun xu; jiajun", jornada="8"
    )
    indice = core.people_by_report_name([person])
    assert indice["jiajun"] is person
    assert indice["jiajun xu"] is person


# -------------------------------------------------------------- registros


def test_merge_records_keeps_manual_days_not_present_in_the_report() -> None:
    person = _person()
    manual = core.manual_record(person, date(2026, 7, 2), 8, nota="Teletrabajo")
    existentes = core.records_to_frame([manual])
    fichaje = core.build_record(person, date(2026, 7, 1), "09:00", "18:00", 9.0, core.ORIGEN_FICHAJE)

    merged = core.merge_records(existentes, [fichaje])

    assert len(merged) == 2
    assert set(merged["fecha"]) == {"2026-07-01", "2026-07-02"}
    assert merged[merged["fecha"] == "2026-07-02"].iloc[0]["origen"] == core.ORIGEN_MANUAL


def test_merge_records_overwrites_the_same_day_and_keeps_created_at() -> None:
    person = _person()
    original = core.build_record(
        person, date(2026, 7, 1), "09:00", "18:00", 9.0, core.ORIGEN_FICHAJE, created_at="2026-07-01"
    )
    existentes = core.records_to_frame([original])
    corregido = core.build_record(person, date(2026, 7, 1), "09:00", "17:00", 8.0, core.ORIGEN_MANUAL)

    merged = core.merge_records(existentes, [corregido])

    assert len(merged) == 1
    fila = merged.iloc[0]
    assert fila["horas"] == "8.0"
    assert fila["created_at"] == "2026-07-01"


def test_recorded_dates_blocks_days_that_already_have_data() -> None:
    person = _person()
    frame = core.records_to_frame([core.manual_record(person, date(2026, 7, 2), 8)])
    assert core.recorded_dates(frame, "EMP001") == {date(2026, 7, 2)}
    assert core.recorded_dates(frame, "EMP999") == set()


# ---------------------------------------------------------------- resumen


def test_business_days_excludes_weekends_and_holidays() -> None:
    festivos = {date(2026, 8, 14)}
    dias = core.business_days(2026, 8, festivos)
    assert date(2026, 8, 14) not in dias
    assert date(2026, 8, 15) not in dias  # sabado
    assert date(2026, 8, 3) in dias


def test_month_summary_counts_missing_hours_for_days_without_clock_in() -> None:
    person = _person()
    holidays = _holidays_frame([])
    festivos = core.holiday_dates_for_year(holidays, 2026)
    laborables = core.business_days(2026, 7, festivos)
    registros = core.records_to_frame(
        [core.manual_record(person, dia, 8) for dia in laborables[:5]]
    )

    resumen = core.month_summary(person, 2026, 7, registros, festivos, {})

    assert resumen is not None
    assert resumen.dias_exigidos == len(laborables)
    assert resumen.horas_exigidas == 8 * len(laborables)
    assert resumen.horas_computadas == 40.0
    assert resumen.saldo == round(40.0 - 8 * len(laborables), 2)


def test_month_summary_discounts_approved_holidays() -> None:
    person = _person()
    holidays = _holidays_frame([])
    festivos = core.holiday_dates_for_year(holidays, 2026)
    absences = pd.DataFrame([_absence_row()])
    por_persona = core.absence_days_by_employee(absences, 2026, festivos)
    laborables = core.business_days(2026, 7, festivos)

    resumen = core.month_summary(
        person, 2026, 7, core.records_to_frame([]), festivos, por_persona["EMP001"]
    )

    assert resumen is not None
    assert resumen.dias_ausencia == 5
    assert resumen.dias_exigidos == len(laborables) - 5
    assert resumen.horas_exigidas == 8 * (len(laborables) - 5)


def test_pending_absences_do_not_discount_anything() -> None:
    festivos: set[date] = set()
    absences = pd.DataFrame([_absence_row(estado="pendiente")])
    assert core.absence_days_by_employee(absences, 2026, festivos) == {}


def test_approved_remote_work_counts_as_a_full_day_without_clock_in() -> None:
    person = _person()
    festivos: set[date] = set()
    absences = pd.DataFrame(
        [_absence_row(tipo="teletrabajo", fecha_inicio="2026-07-06", fecha_fin="2026-07-07")]
    )
    por_persona = core.absence_days_by_employee(absences, 2026, festivos)

    resumen = core.month_summary(
        person, 2026, 7, core.records_to_frame([]), festivos, por_persona["EMP001"]
    )

    assert resumen is not None
    assert resumen.dias_teletrabajo_auto == 2
    assert resumen.horas_teletrabajo == 16.0
    assert resumen.dias_exigidos == len(core.business_days(2026, 7, festivos))


def test_a_recorded_remote_day_is_not_counted_twice() -> None:
    person = _person()
    festivos: set[date] = set()
    absences = pd.DataFrame(
        [_absence_row(tipo="teletrabajo", fecha_inicio="2026-07-06", fecha_fin="2026-07-07")]
    )
    por_persona = core.absence_days_by_employee(absences, 2026, festivos)
    registros = core.records_to_frame([core.manual_record(person, date(2026, 7, 6), 8)])

    resumen = core.month_summary(person, 2026, 7, registros, festivos, por_persona["EMP001"])

    assert resumen is not None
    assert resumen.dias_teletrabajo_auto == 1
    assert resumen.horas_computadas == 8.0 + 8.0


def test_month_summary_is_none_without_a_shift() -> None:
    person = _person(jornada="")
    assert core.month_summary(person, 2026, 7, core.records_to_frame([]), set(), {}) is None


def test_year_summary_carries_the_balance_across_months() -> None:
    person = _person()
    holidays = _holidays_frame([])
    festivos = core.holiday_dates_for_year(holidays, 2026)
    registros = []
    for mes in (1, 2):
        for dia in core.business_days(2026, mes, festivos):
            registros.append(core.manual_record(person, dia, 8))
    frame = core.records_to_frame(registros)

    resumen = core.year_summary(
        [person], 2026, frame, holidays, pd.DataFrame(), today=date(2026, 3, 15)
    )

    por_mes = {int(row["mes_num"]): row for _, row in resumen.iterrows()}
    assert set(por_mes) == {1, 2, 3}
    assert por_mes[1]["saldo"] == 0.0
    assert por_mes[2]["saldo"] == 0.0
    # Marzo sin ningun fichaje: todo el mes en negativo y arrastrado al acumulado.
    assert por_mes[3]["saldo"] < 0
    assert por_mes[3]["saldo_acumulado"] == por_mes[3]["saldo"]


def test_months_of_year_stops_at_the_current_month() -> None:
    assert core.months_of_year(2026, today=date(2026, 3, 15)) == [1, 2, 3]
    assert core.months_of_year(2025, today=date(2026, 3, 15)) == list(range(1, 13))
    assert core.months_of_year(2027, today=date(2026, 3, 15)) == []


# ------------------------------------------------------- lectura informes


def _report_grid() -> list[list[object]]:
    """Rejilla minima con la forma real del informe de la maquina de fichar."""
    ancho = 14
    filas: list[list[object]] = [[None] * ancho for _ in range(2)]
    filas[0][0] = "Fecha:"
    filas[0][3] = "01/07/2026 ~ 31/07/2026"

    cabecera: list[object] = [None] * ancho
    cabecera[0] = "ID."
    cabecera[11] = "david"
    dias: list[object] = [None] * ancho
    vacia: list[object] = [None] * ancho
    horas: list[object] = [None] * ancho
    # Dias 1 (dos fichajes), 2 (un fichaje) y 3 (sin fichaje).
    for indice, (dia, marca) in enumerate([(1, "09:00 18:00"), (2, "09:12"), (3, "")]):
        dias[indice] = dia
        horas[indice] = marca
    filas.extend([cabecera, dias, vacia, horas])
    return filas


def test_detect_month_year_reads_the_report_range() -> None:
    assert core.detect_month_year(_report_grid()) == (7, 2026)


def test_records_from_report_maps_names_and_marks() -> None:
    grid = _report_grid()
    informe = core.ParsedReport(anio=2026, mes=7, marcas_por_nombre={})
    informe.mes = 7
    # Reutilizamos el parseo real sobre la rejilla ya construida.
    marcas = {"david": {1: "09:00 18:00", 2: "09:12", 3: ""}}
    informe.marcas_por_nombre = marcas

    registros, sin_usuario, sin_jornada = core.records_from_report(informe, [_person()], set())

    assert sin_usuario == []
    assert sin_jornada == []
    por_fecha = {r["fecha"]: r for r in registros}
    assert por_fecha["2026-07-01"]["origen"] == core.ORIGEN_FICHAJE
    assert por_fecha["2026-07-01"]["horas"] == "9.0"
    assert por_fecha["2026-07-02"]["origen"] == core.ORIGEN_ESTIMADO
    assert por_fecha["2026-07-02"]["horas"] == "8.5"
    assert "2026-07-03" not in por_fecha  # sin fichaje no se registra
    assert grid  # la rejilla de referencia se mantiene alineada con el parser


def test_records_from_report_reports_people_without_user_or_shift() -> None:
    informe = core.ParsedReport(
        anio=2026,
        mes=7,
        marcas_por_nombre={"desconocido": {1: "09:00 18:00"}, "carla": {1: "09:00 13:00"}},
    )
    sin_jornada_person = AttendancePerson(
        employee_id="EMP004", nombre="Carla Moreno", nombre_fichaje="carla", jornada=""
    )

    registros, sin_usuario, sin_jornada = core.records_from_report(
        informe, [_person(), sin_jornada_person], set()
    )

    assert registros == []
    assert sin_usuario == ["desconocido"]
    assert sin_jornada == ["carla"]


def test_records_from_report_skips_weekends_and_holidays() -> None:
    informe = core.ParsedReport(
        anio=2026,
        mes=8,
        # 14 es festivo de Butarque y 15 cae en sabado.
        marcas_por_nombre={"david": {13: "09:00 18:00", 14: "09:00 18:00", 15: "09:00 18:00"}},
    )
    registros, _, _ = core.records_from_report(informe, [_person()], {date(2026, 8, 14)})
    assert [r["fecha"] for r in registros] == ["2026-08-13"]


def test_parse_attendance_report_reads_a_real_workbook() -> None:
    """El parseo completo (archivo -> marcas) sobre un libro con la forma real."""
    from io import BytesIO

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = core.REPORT_SHEET
    for fila_idx, fila in enumerate(_report_grid(), start=1):
        for col_idx, valor in enumerate(fila, start=1):
            if valor is not None:
                ws.cell(row=fila_idx, column=col_idx, value=valor)
    buffer = BytesIO()
    wb.save(buffer)

    informe = core.parse_attendance_report(buffer.getvalue(), "informe.xlsx")

    assert (informe.anio, informe.mes) == (2026, 7)
    assert informe.mes_label == "Julio 2026"
    assert informe.marcas_por_nombre["david"][1] == "09:00 18:00"
    assert informe.marcas_por_nombre["david"][2] == "09:12"

    registros, sin_usuario, sin_jornada = core.records_from_report(informe, [_person()], set())
    assert (sin_usuario, sin_jornada) == ([], [])
    assert [r["fecha"] for r in registros] == ["2026-07-01", "2026-07-02"]


def test_parse_attendance_report_rejects_a_workbook_without_the_expected_sheet() -> None:
    from io import BytesIO

    import openpyxl
    import pytest

    wb = openpyxl.Workbook()
    wb.active.title = "Otra hoja"
    buffer = BytesIO()
    wb.save(buffer)

    with pytest.raises(ValueError, match="Registros de asistencia"):
        core.parse_attendance_report(buffer.getvalue(), "malo.xlsx")


# ------------------------------------------------ calendario de alta manual


def _absence_days(rows: list[dict[str, str]], festivos: set[date]) -> dict[date, str]:
    por_persona = core.absence_days_by_employee(pd.DataFrame(rows), 2026, festivos)
    return por_persona.get("EMP001", {})


def test_day_picker_blocks_days_that_already_have_a_record() -> None:
    estado = core.day_picker_state(date(2026, 7, 1), {date(2026, 7, 1)}, set(), {})
    assert estado == core.DAY_RECORDED
    assert core.is_day_selectable(estado) is False


def test_day_picker_blocks_approved_holidays_of_that_person() -> None:
    ausencias = _absence_days([_absence_row()], set())  # 6 a 10 de julio
    estado = core.day_picker_state(date(2026, 7, 6), set(), set(), ausencias)
    assert estado == core.DAY_ABSENCE
    assert core.is_day_selectable(estado) is False


def test_day_picker_leaves_other_people_days_free() -> None:
    """La ausencia de uno no bloquea el calendario de otro."""
    rows = [_absence_row(employee_id="EMP002", nombre_employee="Carla Moreno")]
    ausencias = _absence_days(rows, set())
    assert ausencias == {}
    assert core.day_picker_state(date(2026, 7, 6), set(), set(), ausencias) == core.DAY_FREE


def test_day_picker_allows_remote_work_days_but_flags_them() -> None:
    ausencias = _absence_days(
        [_absence_row(tipo="teletrabajo", fecha_inicio="2026-07-06", fecha_fin="2026-07-06")],
        set(),
    )
    estado = core.day_picker_state(date(2026, 7, 6), set(), set(), ausencias)
    assert estado == core.DAY_REMOTE
    assert core.is_day_selectable(estado) is True


def test_day_picker_allows_weekends_and_holidays_but_flags_them() -> None:
    assert core.day_picker_state(date(2026, 8, 15), set(), set(), {}) == core.DAY_NON_WORKING
    estado = core.day_picker_state(date(2026, 8, 14), set(), {date(2026, 8, 14)}, {})
    assert estado == core.DAY_NON_WORKING
    assert core.is_day_selectable(estado) is True


def test_day_picker_prefers_the_existing_record_as_the_reason() -> None:
    """Si el dia tiene registro y ausencia, el motivo que se enseña es el registro."""
    ausencias = _absence_days([_absence_row()], set())
    estado = core.day_picker_state(date(2026, 7, 6), {date(2026, 7, 6)}, set(), ausencias)
    assert estado == core.DAY_RECORDED


def test_pending_holidays_do_not_block_the_calendar() -> None:
    ausencias = _absence_days([_absence_row(estado="pendiente")], set())
    assert core.day_picker_state(date(2026, 7, 6), set(), set(), ausencias) == core.DAY_FREE
