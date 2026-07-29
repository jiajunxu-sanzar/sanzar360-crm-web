from __future__ import annotations

import pandas as pd

from services.history_service import HISTORY_SPECS, HistoryService
from services.riego_campanas import (
    build_riego_timeline_figure,
    export_riego_campanas_excel_bytes,
    normalize_riego_row_values,
    parse_es_nota,
    parse_nota_util,
    serialize_es_nota,
    serialize_nota_util,
    split_riego_and_nota_rows,
    validate_riego_campana_values,
)


def test_riego_campanas_spec_registered() -> None:
    spec = HISTORY_SPECS["riego_campanas"]
    assert spec.worksheet_name == "HistoricoRiegos_Campanas"
    assert spec.id_column == "historial_riego_id"
    assert spec.date_column == "dia_riego"
    for col in (
        "historial_campana_id",
        "hora_inicio_riego",
        "horas_riego",
        "litros",
        "es_nota",
        "nota",
        "nota_util",
    ):
        assert col in spec.headers


def test_parse_es_nota_variants() -> None:
    assert parse_es_nota("true") is True
    assert parse_es_nota("TRUE") is True
    assert parse_es_nota("sí") is True
    assert parse_es_nota("si") is True
    assert parse_es_nota("1") is True
    assert parse_es_nota("false") is False
    assert parse_es_nota("") is False
    assert parse_es_nota("0") is False
    assert serialize_es_nota(True) == "true"
    assert serialize_es_nota(False) == "false"


def test_parse_nota_util_variants() -> None:
    assert parse_nota_util("") is True
    assert parse_nota_util("true") is True
    assert parse_nota_util("false") is False
    assert parse_nota_util("no") is False
    assert serialize_nota_util(True) == "true"
    assert serialize_nota_util(False) == "false"


def test_validate_nota_without_day_or_time() -> None:
    values = {
        "es_nota": "true",
        "dia_riego": "",
        "hora_inicio_riego": "",
        "nota": "",
        "horas_riego": "",
        "litros": "",
        "nota_util": "true",
    }
    assert validate_riego_campana_values(values) is not None
    values["nota"] = "Revisar caudal"
    assert validate_riego_campana_values(values) is None


def test_validate_riego_requires_numeric_fields() -> None:
    values = {
        "es_nota": "false",
        "dia_riego": "10/06/2026",
        "hora_inicio_riego": "08:30",
        "horas_riego": "",
        "litros": "100",
        "nota": "",
    }
    assert validate_riego_campana_values(values) is not None
    values["horas_riego"] = "2,5"
    assert validate_riego_campana_values(values) is None


def test_normalize_nota_clears_riego_fields() -> None:
    out = normalize_riego_row_values(
        {
            "es_nota": "sí",
            "dia_riego": "2026-06-10",
            "hora_inicio_riego": "9:05",
            "horas_riego": "2",
            "litros": "10",
            "nota": "hola",
            "nota_util": "false",
        }
    )
    assert out["es_nota"] == "true"
    assert out["dia_riego"] == ""
    assert out["hora_inicio_riego"] == ""
    assert out["horas_riego"] == ""
    assert out["litros"] == ""
    assert out["nota"] == "hola"
    assert out["nota_util"] == "false"


def test_normalize_riego_sets_nota_util_false() -> None:
    out = normalize_riego_row_values(
        {
            "es_nota": "false",
            "dia_riego": "10/06/2026",
            "hora_inicio_riego": "08:00",
            "horas_riego": "1",
            "litros": "50",
            "nota": "",
            "nota_util": "true",
        }
    )
    assert out["nota_util"] == "false"


class _FakeSheets:
    def __init__(self) -> None:
        headers = list(HISTORY_SPECS["riego_campanas"].headers)
        self.frames = {"HistoricoRiegos_Campanas": pd.DataFrame(columns=headers)}

    def get_or_create_worksheet(self, name: str, headers: list[str]) -> None:
        if name not in self.frames:
            self.frames[name] = pd.DataFrame(columns=headers)

    def read_worksheet_df(self, name: str, headers: list[str]) -> pd.DataFrame:
        df = self.frames.get(name, pd.DataFrame(columns=headers)).copy()
        for h in headers:
            if h not in df.columns:
                df[h] = ""
        return df[headers].fillna("").astype(str)

    def row_numbers_by_id(self, name: str, id_header: str) -> dict[str, int]:
        df = self.frames[name].fillna("").astype(str)
        return {
            str(row[id_header]).strip(): i + 2
            for i, row in df.iterrows()
            if str(row.get(id_header, "")).strip()
        }

    def append_worksheet_row(self, name: str, headers: list[str], row: dict) -> None:
        if name not in self.frames:
            self.frames[name] = pd.DataFrame(columns=headers)
        new_row = {h: str(row.get(h, "") or "") for h in headers}
        self.frames[name] = pd.concat([self.frames[name], pd.DataFrame([new_row])], ignore_index=True)

    def update_worksheet_row(self, name: str, headers: list[str], row_num: int, row: dict) -> None:
        idx = max(0, row_num - 2)
        df = self.frames[name].copy()
        for h in headers:
            df.at[idx, h] = str(row.get(h, "") or "")
        self.frames[name] = df

    def write_worksheet_df(self, name: str, df: pd.DataFrame, headers: list[str]) -> None:
        self.frames[name] = df[headers].fillna("").astype(str).copy()

    def delete_rows_where_column_equals(self, worksheet_title: str, column_name: str, value: str) -> int:
        if worksheet_title not in self.frames:
            return 0
        df = self.frames[worksheet_title].copy()
        if column_name not in df.columns:
            return 0
        mask = df[column_name].astype(str).str.strip() == str(value).strip()
        removed = int(mask.sum())
        self.frames[worksheet_title] = df.loc[~mask].copy()
        return removed


def test_rows_for_campaign_filters_and_orders() -> None:
    sheets = _FakeSheets()
    headers = list(HISTORY_SPECS["riego_campanas"].headers)
    sheets.frames["HistoricoRiegos_Campanas"] = pd.DataFrame(
        [
            {
                "historial_riego_id": "r2",
                "historial_campana_id": "camp-a",
                "contact_id": "c1",
                "dia_riego": "11/06/2026",
                "hora_inicio_riego": "07:00",
                "horas_riego": "1",
                "litros": "10",
                "es_nota": "false",
                "nota": "",
                "nota_util": "false",
                "created_at": "",
                "updated_at": "",
            },
            {
                "historial_riego_id": "r1",
                "historial_campana_id": "camp-a",
                "contact_id": "c1",
                "dia_riego": "10/06/2026",
                "hora_inicio_riego": "18:00",
                "horas_riego": "2",
                "litros": "20",
                "es_nota": "false",
                "nota": "",
                "nota_util": "false",
                "created_at": "",
                "updated_at": "",
            },
            {
                "historial_riego_id": "r0",
                "historial_campana_id": "camp-a",
                "contact_id": "c1",
                "dia_riego": "10/06/2026",
                "hora_inicio_riego": "08:00",
                "horas_riego": "1",
                "litros": "5",
                "es_nota": "false",
                "nota": "",
                "nota_util": "false",
                "created_at": "",
                "updated_at": "",
            },
            {
                "historial_riego_id": "rx",
                "historial_campana_id": "camp-b",
                "contact_id": "c1",
                "dia_riego": "09/06/2026",
                "hora_inicio_riego": "08:00",
                "horas_riego": "1",
                "litros": "5",
                "es_nota": "false",
                "nota": "",
                "nota_util": "false",
                "created_at": "",
                "updated_at": "",
            },
        ],
        columns=headers,
    ).fillna("").astype(str)

    svc = HistoryService(sheets)  # type: ignore[arg-type]
    ordered = svc.rows_for_campaign("riego_campanas", "camp-a")
    assert [r["historial_riego_id"] for r in ordered] == ["r0", "r1", "r2"]
    assert svc.rows_for_campaign("riego_campanas", "missing") == []


def test_riego_campanas_add_update_delete_with_nota_util() -> None:
    sheets = _FakeSheets()
    svc = HistoryService(sheets)  # type: ignore[arg-type]
    created = svc.add_row(
        "riego_campanas",
        {
            "historial_campana_id": "camp-1",
            "contact_id": "c-1",
            "dia_riego": "",
            "hora_inicio_riego": "",
            "horas_riego": "",
            "litros": "",
            "es_nota": "true",
            "nota": "Observación",
            "nota_util": "true",
        },
    )
    riego_id = created["historial_riego_id"]
    rows = svc.rows_for_campaign("riego_campanas", "camp-1")
    assert len(rows) == 1
    assert rows[0]["nota"] == "Observación"
    assert rows[0]["nota_util"] == "true"

    svc.update_row(
        "riego_campanas",
        riego_id,
        {
            "historial_campana_id": "camp-1",
            "contact_id": "c-1",
            "es_nota": "true",
            "nota": "Obsoleta",
            "nota_util": "false",
        },
    )
    rows = svc.rows_for_campaign("riego_campanas", "camp-1")
    assert rows[0]["nota_util"] == "false"
    assert rows[0]["nota"] == "Obsoleta"

    svc.delete_row("riego_campanas", riego_id)
    assert svc.rows_for_campaign("riego_campanas", "camp-1") == []


def test_build_riego_timeline_excludes_notas() -> None:
    fig = build_riego_timeline_figure(
        [
            {
                "dia_riego": "12/06/2026",
                "hora_inicio_riego": "08:00",
                "horas_riego": "2",
                "litros": "100",
                "es_nota": "false",
                "nota": "",
            },
            {
                "dia_riego": "",
                "hora_inicio_riego": "",
                "horas_riego": "",
                "litros": "",
                "es_nota": "true",
                "nota": "Parada por viento",
                "nota_util": "true",
            },
        ]
    )
    assert fig is not None
    assert len(fig.data) == 1


def test_build_riego_timeline_empty_when_only_notas() -> None:
    fig = build_riego_timeline_figure(
        [
            {
                "es_nota": "true",
                "nota": "Solo nota",
                "nota_util": "false",
            }
        ]
    )
    assert fig is None


def test_split_riego_and_nota_rows() -> None:
    riegos, notas = split_riego_and_nota_rows(
        [
            {"historial_riego_id": "1", "es_nota": "false"},
            {"historial_riego_id": "2", "es_nota": "true"},
            {"historial_riego_id": "3", "es_nota": "sí"},
        ]
    )
    assert [r["historial_riego_id"] for r in riegos] == ["1"]
    assert [r["historial_riego_id"] for r in notas] == ["2", "3"]


def test_export_riego_campanas_excel_separates_notas() -> None:
    from io import BytesIO

    # 1 riego + 1 nota: deben acabar en hojas distintas.
    rows = [
        {
            "historial_riego_id": "r1",
            "historial_campana_id": "camp-1",
            "contact_id": "c-1",
            "dia_riego": "10/06/2026",
            "hora_inicio_riego": "08:00",
            "horas_riego": "1",
            "litros": "20",
            "es_nota": "false",
            "nota": "",
            "nota_util": "false",
            "created_at": "29/07/2026",
            "updated_at": "29/07/2026",
        },
        {
            "historial_riego_id": "n1",
            "historial_campana_id": "camp-1",
            "contact_id": "c-1",
            "dia_riego": "",
            "hora_inicio_riego": "",
            "horas_riego": "",
            "litros": "",
            "es_nota": "true",
            "nota": "Parada por viento",
            "nota_util": "true",
            "created_at": "29/07/2026",
            "updated_at": "29/07/2026",
        },
    ]

    xlsx = export_riego_campanas_excel_bytes(rows)
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(xlsx), data_only=True)
    assert set(wb.sheetnames) == {"Riegos", "Notas"}

    ws_r = wb["Riegos"]
    data_r = list(ws_r.iter_rows(min_row=2, values_only=True))
    assert len(data_r) == 1
    # Columna 0 = dia_riego (orden fijo por DataFrame).
    assert str(data_r[0][0]) == "10/06/2026"

    ws_n = wb["Notas"]
    data_n = list(ws_n.iter_rows(min_row=2, values_only=True))
    assert len(data_n) == 1
    # Columna 0 = nota, col 1 = nota_util.
    assert str(data_n[0][0]) == "Parada por viento"
    assert str(data_n[0][1]).strip() == "Útil"
