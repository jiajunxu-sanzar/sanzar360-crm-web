"""Tests for full contact ficha Excel/PDF export."""
from __future__ import annotations

from datetime import datetime
from io import BytesIO

from services.contact_ficha_export import (
    EXPORT_HISTORY_KINDS,
    SHEET_TITLES,
    build_contact_ficha_pdf_bytes,
    build_contact_ficha_xlsx_bytes,
)
from services.history_service import HISTORY_SPECS

_EXPORTED_AT = datetime(2026, 8, 24, 11, 0)


def _contact() -> dict[str, str]:
    return {
        "contact_id": "cid-demo-001",
        "nombre": "Finca La Vega",
        "estado": "Cliente",
        "municipio": "Murcia",
        "correo": "demo@example.com",
        "telefono": "",  # vacío: no debe salir en ficha
    }


def _histories() -> dict[str, list[dict[str, str]]]:
    return {
        "seguimiento_comercial": [
            {
                "historial_accion_id": "a1",
                "contact_id": "cid-demo-001",
                "nombre_cliente": "Finca La Vega",
                "fecha_contacto": "01/08/2026",
                "hora_contacto": "10:00",
                "persona_contacto": "David",
                "canal_contacto": "llamada",
                "resultado_contacto": "exitoso",
                "notas_contacto": "Todo bien",
            }
        ],
        "tareas": [],
        "notas": [
            {
                "historial_nota_id": "n1",
                "contact_id": "cid-demo-001",
                "titulo": "Nota útil",
                "tipo_nota": "general",
                "estado_nota": "Útil",
                "notas": "Detalle largo",
                "persona_nota": "David",
                "fecha_creacion": "01/08/2026",
                "fecha_update": "01/08/2026",
            }
        ],
        "sensores": [],
        "campanas": [],
        "suscripciones": [],
        "incidencias": [],
    }


def test_build_contact_ficha_xlsx_has_expected_sheets() -> None:
    content, fname = build_contact_ficha_xlsx_bytes(
        _contact(), _histories(), exported_at=_EXPORTED_AT
    )
    assert fname.endswith(".xlsx")
    assert "finca_la_vega" in fname
    assert "20260824" in fname

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(content))
    expected = {"Indice", "Ficha", *[SHEET_TITLES[k] for k in EXPORT_HISTORY_KINDS]}
    assert expected.issubset(set(wb.sheetnames))

    ws_ficha = wb["Ficha"]
    campos = [ws_ficha.cell(row=r, column=1).value for r in range(2, ws_ficha.max_row + 1)]
    assert "contact_id" in campos
    assert "nombre" in campos
    assert "correo" in campos
    assert "telefono" not in campos  # vacío

    ws_seg = wb["Seguimiento"]
    assert ws_seg.cell(row=2, column=1).value == "a1"

    ws_tareas = wb["Tareas"]
    assert ws_tareas.cell(row=2, column=1).value == "Sin registros"


def test_build_contact_ficha_pdf_returns_pdf() -> None:
    content, fname = build_contact_ficha_pdf_bytes(
        _contact(), _histories(), exported_at=_EXPORTED_AT
    )
    assert fname.endswith(".pdf")
    assert content[:4] == b"%PDF"
    assert len(content) > 100


def test_build_contact_ficha_empty_histories_ok() -> None:
    empty = {k: [] for k in EXPORT_HISTORY_KINDS}
    xlsx, _ = build_contact_ficha_xlsx_bytes(_contact(), empty, exported_at=_EXPORTED_AT)
    pdf, _ = build_contact_ficha_pdf_bytes(_contact(), empty, exported_at=_EXPORTED_AT)
    assert xlsx
    assert pdf[:4] == b"%PDF"
    # Todas las hojas de histórico existen
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(xlsx))
    for kind in EXPORT_HISTORY_KINDS:
        assert SHEET_TITLES[kind] in wb.sheetnames
        assert len(HISTORY_SPECS[kind].headers) >= 1
